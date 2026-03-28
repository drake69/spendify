"""Deterministic normalization pipeline (RF-02, RF-03, RF-04, RF-06).

All functions are pure / side-effect-free and unit-testable without LLM mocks.
"""
from __future__ import annotations

import hashlib
import math
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
from itertools import combinations, permutations as _permutations
from statistics import median
from typing import Optional

import chardet
import pandas as pd

from core.models import (
    Confidence,
    DocumentType,
    GirocontoMode,
    SignConvention,
    TransactionType,
)
from core.schemas import DocumentSchema
from support.logging import setup_logging

logger = setup_logging()


# ── Pre-processing constants ───────────────────────────────────────────────────

_PREHEADER_MAX_ROWS: int = 20        # absolute max rows we are willing to strip
_PREHEADER_MAX_RATIO: float = 0.10   # 10 % of total rows — safety cap
_PREHEADER_DENSITY_THRESHOLD: float = 0.5   # fraction of median density below which a row is "sparse"
_LOW_VARIABILITY_RATIO: float = 0.015  # nunique/nrows < 1.5 % → metadata/constant column


@dataclass
class PreprocessInfo:
    """Metadata produced by the Phase-0 preprocessing step.

    Carried alongside the raw DataFrame so the rest of the pipeline can log
    or display what was stripped/dropped without re-running the analysis.
    """
    skipped_rows: int = 0
    footer_rows_stripped: int = 0
    footer_strip_method: str = ""
    """How footer rows were stripped: 'phase1', 'phase1+phase2', 'phase1+phase3', 'iqr_fallback', etc."""
    footer_patterns_learned: int = 0
    """Number of new footer patterns learned from Phase 2 LLM extraction."""
    dropped_columns: list[str] = field(default_factory=list)
    columns_before_drop: list[str] = field(default_factory=list)
    """Column names captured AFTER preheader-strip but BEFORE drop_low_variability_columns.
    Used for stable schema lookup (cols_key) regardless of file size."""
    header_certain: bool = False
    """Whether the pre-load header detection was certain (used for confidence scoring)."""
    border_detected: bool = False
    """Whether the table bounds were determined by Excel border detection."""
    border_region: Optional[tuple[int, int, int, int]] = None
    """(first_row, last_row, first_col, last_col) 0-based. None if no borders."""


# ── Encoding / format detection ───────────────────────────────────────────────

def detect_encoding(raw_bytes: bytes) -> str:
    """Detect file encoding using chardet with fallback heuristics (I-08).

    If chardet confidence is low (< 0.70), tries common European encodings
    (latin1, cp1252, utf-8) and picks the one with fewest replacement characters.
    """
    result = chardet.detect(raw_bytes)
    enc = (result.get("encoding") or "utf-8").lower()
    confidence = result.get("confidence", 0.0)

    # Normalize common aliases
    if enc in ("ascii",):
        enc = "utf-8"

    # I-08: fallback when chardet is uncertain
    if confidence < 0.70:
        logger.info(
            f"detect_encoding: chardet returned {enc} with low confidence "
            f"({confidence:.2f}) — trying European fallbacks"
        )
        # Try common European encodings; pick the one with fewest decode errors
        candidates = ["utf-8", "latin-1", "cp1252", enc]
        best_enc = enc
        best_errors = float("inf")
        sample = raw_bytes[:8192]  # check first 8KB only

        for candidate in candidates:
            try:
                decoded = sample.decode(candidate, errors="replace")
                # Count replacement characters (U+FFFD)
                n_errors = decoded.count("\ufffd")
                if n_errors < best_errors:
                    best_errors = n_errors
                    best_enc = candidate
            except (LookupError, UnicodeDecodeError):
                continue

        if best_enc != enc:
            logger.info(
                f"detect_encoding: fallback selected {best_enc} "
                f"(was {enc}, errors: {best_errors} vs original)"
            )
        enc = best_enc

    return enc


def detect_delimiter(content: str) -> str:
    """Detect CSV delimiter by frequency analysis."""
    candidates = [",", ";", "\t", "|"]
    counts = {d: content.count(d) for d in candidates}
    return max(counts, key=counts.get)


def _row_density(fields: list[str]) -> float:
    """Fraction of non-empty fields in a row."""
    if not fields:
        return 0.0
    return sum(1 for f in fields if f.strip()) / len(fields)


def _row_non_numeric_count(fields: list[str]) -> int:
    """Count non-empty, non-numeric fields (text = likely column names)."""
    return sum(
        1 for f in fields
        if f.strip() and not re.match(r'^[\d\.\,\-\+\s€$£%]+$', f.strip())
    )


def detect_header_row(lines: list[str]) -> tuple[int, bool]:
    """Detect the header row using density analysis.

    Algorithm:
    1. Parse each line into fields.
    2. The header row is the one with the highest density (most non-empty cells).
    3. Tiebreak: prefer the row with more non-numeric fields (text = column names).
    4. certain=True if the winner has density ≥ 0.5 AND ≥ 2 non-numeric fields.

    Returns (index, certain).
    """
    if not lines:
        return 0, False

    max_scan = min(len(lines), 50)
    best_idx = 0
    best_density = 0.0
    best_text_count = 0

    logger.info("── detect_header_row: scanning %d lines ──", max_scan)
    for i in range(max_scan):
        fields = [f.strip() for f in re.split(r'[,;\t|]', lines[i])]
        density = _row_density(fields)
        text_count = _row_non_numeric_count(fields)
        n_fields = len([f for f in fields if f])

        logger.info(
            "  row %2d | density=%.2f | text_fields=%d | filled=%d/%d | preview=%.80s",
            i, density, text_count, n_fields, len(fields),
            lines[i][:80].replace("\n", ""),
        )

        # Prefer higher density; tiebreak by more non-numeric (text) fields
        if (density > best_density) or (density == best_density and text_count > best_text_count):
            best_idx = i
            best_density = density
            best_text_count = text_count

    certain = best_density >= 0.5 and best_text_count >= 2
    logger.info(
        "── detect_header_row result: row=%d density=%.2f text_count=%d certain=%s ──",
        best_idx, best_density, best_text_count, certain,
    )
    return best_idx, certain


def _cell_has_border(cell) -> bool:
    """Check if a cell has any non-trivial border on any side."""
    border = getattr(cell, "border", None)
    if border is None:
        return False
    for side_name in ("left", "right", "top", "bottom"):
        side = getattr(border, side_name, None)
        if side is not None and getattr(side, "style", None) not in (None, "none"):
            return True
    return False


def detect_bordered_region(
    raw_bytes: bytes,
    filename: str = "",
    max_scan_rows: int = 60,
) -> Optional[tuple[int, int, int, int]]:
    """Detect a bordered table rectangle in an Excel file.

    Scans the first *max_scan_rows* rows for a contiguous rectangular region
    where ALL cells have at least one border (thin/medium/thick on any side).

    Returns ``(first_row, last_row, first_col, last_col)`` as **0-based**
    indices, or ``None`` if no qualifying bordered region is found.

    Only works on ``.xlsx`` files — ``.xls`` always returns ``None``.
    """
    # Only .xlsx — openpyxl doesn't expose borders for .xls
    if filename and not filename.lower().endswith(".xlsx"):
        return None

    import io as _io
    try:
        import openpyxl
        wb = openpyxl.load_workbook(
            _io.BytesIO(raw_bytes), read_only=False, data_only=True,
        )
        sheet_name = detect_best_sheet(wb)
        ws = wb[sheet_name]

        # ── Build border grid ──────────────────────────────────────
        max_col = ws.max_column or 0
        max_row = min(ws.max_row or 0, max_scan_rows)
        if max_row < 3 or max_col < 3:
            wb.close()
            return None

        # has_border[r][c] — 0-based
        grid: list[list[bool]] = []
        for r_idx in range(1, max_row + 1):
            row_borders: list[bool] = []
            for c_idx in range(1, max_col + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                row_borders.append(_cell_has_border(cell))
            grid.append(row_borders)

        wb.close()

        # ── Find largest bordered rectangle (sweep top-down) ──────
        # For each row, compute the longest contiguous run of bordered cells.
        def _bordered_run(row_bools: list[bool]) -> Optional[tuple[int, int]]:
            """Return (start, end) of the longest contiguous True run, or None."""
            best_start = best_end = -1
            best_len = 0
            start = None
            for i, v in enumerate(row_bools):
                if v:
                    if start is None:
                        start = i
                else:
                    if start is not None:
                        run_len = i - start
                        if run_len > best_len:
                            best_start, best_end, best_len = start, i - 1, run_len
                        start = None
            # Trailing run
            if start is not None:
                run_len = len(row_bools) - start
                if run_len > best_len:
                    best_start, best_end = start, len(row_bools) - 1
            if best_start < 0:
                return None
            return best_start, best_end

        # Find first row with a bordered run >= 3 columns
        region_r1: Optional[int] = None
        region_c1: int = 0
        region_c2: int = 0

        for r, row_bools in enumerate(grid):
            run = _bordered_run(row_bools)
            if run and (run[1] - run[0] + 1) >= 3:
                region_r1 = r
                region_c1, region_c2 = run
                break

        if region_r1 is None:
            return None

        # Extend downward: intersect column range
        region_r2 = region_r1
        for r in range(region_r1 + 1, len(grid)):
            run = _bordered_run(grid[r])
            if run is None:
                break
            # Intersect with current column range
            new_c1 = max(region_c1, run[0])
            new_c2 = min(region_c2, run[1])
            if new_c2 - new_c1 + 1 < 3:
                break  # intersection too narrow
            region_c1, region_c2 = new_c1, new_c2
            region_r2 = r

        # Validate: minimum 3 rows × 3 columns
        if (region_r2 - region_r1 + 1) < 3:
            return None

        logger.info(
            "── detect_bordered_region: found region rows=%d..%d cols=%d..%d (%d×%d) ──",
            region_r1, region_r2, region_c1, region_c2,
            region_r2 - region_r1 + 1, region_c2 - region_c1 + 1,
        )
        return (region_r1, region_r2, region_c1, region_c2)

    except Exception as exc:
        logger.warning("detect_bordered_region failed: %s", exc)
        return None


def detect_header_row_excel(
    raw_bytes: bytes,
    filename: str = "",
    skip_border_check: bool = False,
) -> tuple[int, bool, Optional[tuple[int, int, int, int]]]:
    """Detect header row in an Excel file.

    Strategy:
    1. **Border detection first** (deterministic): if a bordered region is found,
       the first row with max density inside the region = header. certain=True.
    2. **Density fallback**: scan first 50 rows for highest density row.

    Args:
        skip_border_check: If True, skip border detection (I-10 optimisation for
            formats known to have no borders from cached schema ``has_borders=False``).

    Returns ``(skip_rows, certain, border_region)``.
    ``border_region`` is ``(r1, r2, c1, c2)`` 0-based, or ``None``.
    """
    import io as _io

    # ── Phase A: Try border detection ──────────────────────────────
    if skip_border_check:
        logger.info("── detect_header_row_excel: skipping border check (has_borders=False from cache) ──")
        border_region = None
    else:
        border_region = detect_bordered_region(raw_bytes, filename)
    if border_region is not None:
        r1, r2, c1, c2 = border_region
        logger.info(
            "── detect_header_row_excel: bordered region rows=%d..%d cols=%d..%d → scanning for header ──",
            r1, r2, c1, c2,
        )
        try:
            import openpyxl
            wb = openpyxl.load_workbook(_io.BytesIO(raw_bytes), read_only=True, data_only=True)
            sheet_name = detect_best_sheet(wb)
            ws = wb[sheet_name]

            best_idx = r1
            best_density = 0.0
            best_text_count = 0

            for i, row in enumerate(ws.iter_rows(values_only=True, max_row=r2 + 1)):
                if i < r1:
                    continue
                if i > r2:
                    break
                # Restrict to bordered columns
                all_fields = [str(v).strip() if v is not None else "" for v in row]
                fields = all_fields[c1:c2 + 1]
                density = _row_density(fields)
                text_count = _row_non_numeric_count(fields)

                logger.info(
                    "  [border] row %2d | density=%.2f | text=%d | values=%s",
                    i, density, text_count, [f[:30] for f in fields if f][:6],
                )

                # Cross-check: first row with 100% density = header
                if density >= 1.0 and text_count >= 2:
                    best_idx = i
                    best_density = density
                    best_text_count = text_count
                    break  # 100% density = certain header, stop

                if (density > best_density) or (density == best_density and text_count > best_text_count):
                    best_idx = i
                    best_density = density
                    best_text_count = text_count

            wb.close()
            logger.info(
                "── detect_header_row_excel (border): row=%d density=%.2f text=%d certain=True ──",
                best_idx, best_density, best_text_count,
            )
            return best_idx, True, border_region

        except Exception as exc:
            logger.warning("detect_header_row_excel border scan failed: %s — falling back to density", exc)

    # ── Phase B: Density-only fallback ─────────────────────────────
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_io.BytesIO(raw_bytes), read_only=True, data_only=True)
        sheet_name = detect_best_sheet(wb)
        ws = wb[sheet_name]

        best_idx = 0
        best_density = 0.0
        best_text_count = 0

        logger.info("── detect_header_row_excel (density): sheet='%s' ──", sheet_name)
        for i, row in enumerate(ws.iter_rows(values_only=True, max_row=50)):
            fields = [str(v).strip() if v is not None else "" for v in row]
            density = _row_density(fields)
            text_count = _row_non_numeric_count(fields)
            n_fields = len([f for f in fields if f])

            logger.info(
                "  row %2d | density=%.2f | text_fields=%d | filled=%d/%d | values=%s",
                i, density, text_count, n_fields, len(fields),
                [f[:30] for f in fields if f][:6],
            )

            if (density > best_density) or (density == best_density and text_count > best_text_count):
                best_idx = i
                best_density = density
                best_text_count = text_count

        wb.close()
        certain = best_density >= 0.5 and best_text_count >= 2
        logger.info(
            "── detect_header_row_excel result: row=%d density=%.2f text_count=%d certain=%s ──",
            best_idx, best_density, best_text_count, certain,
        )
        return best_idx, certain, None
    except Exception as exc:
        logger.warning("detect_header_row_excel failed: %s", exc)
    return 0, False, None


def detect_skip_rows(
    raw_bytes: bytes,
    filename: str,
    skip_border_check: bool = False,
) -> tuple[int, bool, Optional[tuple[int, int, int, int]]]:
    """Unified pre-load header detection for CSV and Excel.

    Returns ``(skip_rows, certain, border_region)``:
      certain=True  → detection confident, use skip_rows silently.
      certain=False → could not determine; surface a number_input to the user.
      border_region → (r1, r2, c1, c2) 0-based if Excel borders found, else None.

    Args:
        skip_border_check: If True, skip border detection entirely (I-10 optimisation
            for formats known to have no borders from cached schema ``has_borders=False``).
    """
    logger.info("══ detect_skip_rows [%s] (%d bytes, type=%s, skip_border=%s) ══",
                filename, len(raw_bytes),
                "excel" if filename.lower().endswith((".xlsx", ".xls")) else "csv",
                skip_border_check)
    border_region: Optional[tuple[int, int, int, int]] = None
    if filename.lower().endswith((".xlsx", ".xls")):
        skip, certain, border_region = detect_header_row_excel(
            raw_bytes, filename, skip_border_check=skip_border_check,
        )
    else:
        encoding = detect_encoding(raw_bytes)
        text = raw_bytes.decode(encoding, errors="replace")
        skip, certain = detect_header_row(text.splitlines())
    logger.info(
        "══ detect_skip_rows result: skip_rows=%d, certain=%s, border=%s ══",
        skip, certain, border_region,
    )
    return skip, certain, border_region


def detect_best_sheet(workbook) -> str:
    """Select the sheet with the most numeric columns and rows,
    excluding sheets whose name matches summary patterns."""
    summary_re = re.compile(r'summary|totale|riepilogo', re.IGNORECASE)
    best_sheet = None
    best_score = -1

    for name in workbook.sheetnames:
        if summary_re.search(name):
            continue
        ws = workbook[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        n_rows = len(rows)
        n_numeric_cols = sum(
            1 for col_idx in range(len(rows[0]))
            if sum(1 for row in rows if _is_numeric_cell(row[col_idx] if col_idx < len(row) else None)) > n_rows * 0.5
        )
        score = n_rows + n_numeric_cols * 10
        if score > best_score:
            best_score = score
            best_sheet = name

    return best_sheet or workbook.sheetnames[0]


def _is_numeric_cell(value) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    try:
        float(str(value).replace(",", ".").replace(" ", ""))
        return True
    except ValueError:
        return False


# ── Date normalization ────────────────────────────────────────────────────────

_DATE_FORMATS_FALLBACK = [
    "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y",
    "%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%m/%d/%y",
]


def parse_date_safe(value: str, fmt: str) -> Optional[date]:
    """Parse a date string. Tries fmt first, then common Italian/ISO fallbacks."""
    if not value or not isinstance(value, str):
        return None
    import datetime
    v = value.strip()
    # Try the specified format first
    if fmt:
        try:
            return datetime.datetime.strptime(v, fmt).date()
        except ValueError:
            pass
    # Fallback: try common formats
    for fallback in _DATE_FORMATS_FALLBACK:
        if fallback == fmt:
            continue
        try:
            return datetime.datetime.strptime(v, fallback).date()
        except ValueError:
            continue
    return None


# ── Amount normalization ──────────────────────────────────────────────────────

def parse_amount(value: str | float | int | Decimal) -> Optional[Decimal]:
    """Convert a raw amount value to Decimal. Returns None on failure or non-finite values."""
    if isinstance(value, Decimal):
        return value if value.is_finite() else None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not math.isfinite(value):
            return None
        return Decimal(str(value))
    if not isinstance(value, str):
        return None
    s = str(value).strip()
    s = re.sub(r'[€$£\s]', '', s)
    # Detect separators
    has_dot = '.' in s
    has_comma = ',' in s
    if has_dot and has_comma:
        # e.g. "1.234,56" → European; "1,234.56" → US
        if s.rfind('.') < s.rfind(','):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif has_comma:
        # Could be decimal separator
        parts = s.split(',')
        if len(parts) == 2 and len(parts[1]) <= 2:
            s = s.replace(',', '.')
        else:
            s = s.replace(',', '')
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def apply_sign_convention(
    row: dict,
    amount_col: str,
    debit_col: Optional[str],
    credit_col: Optional[str],
    convention: SignConvention,
) -> Optional[Decimal]:
    """Return a signed Decimal for the transaction amount (negative = expense)."""
    if convention == SignConvention.signed_single:
        return parse_amount(row.get(amount_col))
    elif convention == SignConvention.debit_positive:
        debit = parse_amount(row.get(debit_col)) if debit_col else None
        credit = parse_amount(row.get(credit_col)) if credit_col else None
        # If neither column parsed, fall back to amount_col (schema may have mismapped)
        if debit is None and credit is None:
            return parse_amount(row.get(amount_col)) if amount_col else None
        # Use abs() so the result is correct whether the bank stores values as
        # positive (standard) or negative (e.g. YELLOW "Uscite" = -2.60).
        # Column NAME determines direction; the sign in the cell is irrelevant.
        debit_abs = abs(debit) if debit is not None else Decimal(0)
        credit_abs = abs(credit) if credit is not None else Decimal(0)
        return credit_abs - debit_abs
    elif convention == SignConvention.credit_negative:
        # credit column is positive (income), debit column is negative (expense)
        credit = parse_amount(row.get(credit_col)) if credit_col else None
        debit = parse_amount(row.get(debit_col)) if debit_col else None
        if credit is None and debit is None:
            return parse_amount(row.get(amount_col)) if amount_col else None
        if credit and credit > 0:
            return credit
        if debit:
            return -abs(debit)
        return parse_amount(row.get(amount_col))
    return None


# ── Description normalization ─────────────────────────────────────────────────

def normalize_description(text: str) -> str:
    """casefold + trim + unicode NFC"""
    if not text:
        return ""
    text = unicodedata.normalize("NFC", str(text))
    return text.casefold().strip()


# ── SHA-256 idempotency key ────────────────────────────────────────────────────

def compute_transaction_id(source_file: str, raw_date: str, raw_amount: str, raw_description: str,
                           account_label: str = "") -> str:
    """Return the first 24 chars of SHA-256 of the raw (pre-normalisation) fields.

    Using raw strings keeps the key stable across normalisation changes:
    improving parse_amount or normalize_description won't shift existing IDs.
    For debit/credit conventions pass raw_amount as '<debit_raw>|<credit_raw>'.

    Deduplication key uses account_label (stable bank account identifier) instead
    of source_file (filename), so the same transaction imported from two different
    files of the same account is correctly recognised as a duplicate.
    Falls back to source_file if account_label is empty (e.g. unclassified files).
    """
    account_key = account_label.strip() if account_label and account_label.strip() else source_file
    key = f"{account_key}|{raw_date}|{raw_amount}|{raw_description}"
    return hashlib.sha256(key.encode("utf-8")).hexdigest()[:24]


def compute_columns_key(df: "pd.DataFrame") -> str:  # type: ignore[name-defined]
    """Return SHA-256[:16] of the sorted column names.

    Used as the DocumentSchema cache key so the same bank layout is recognised
    regardless of the export filename (e.g. CARTA_2025.xlsx vs CARTA_2026.xlsx).
    Columns are sorted to be robust to minor reordering.
    """
    cols = "|".join(sorted(str(c).strip() for c in df.columns))
    return "cols:" + hashlib.sha256(cols.encode("utf-8")).hexdigest()[:16]


def compute_file_hash(raw_bytes: bytes) -> str:
    """Return SHA-256 hex of raw file bytes (for import-level idempotency)."""
    return hashlib.sha256(raw_bytes).hexdigest()


def compute_header_sha256(raw_bytes: bytes, filename: str, n: int = 30) -> str:
    """SHA256 of the first min(n, total) raw rows — stable fingerprint of the file header/pre-header area."""
    import io as _io
    name_lower = filename.lower()
    if name_lower.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(_io.BytesIO(raw_bytes), read_only=True, data_only=True)
            sheet_name = detect_best_sheet(wb)
            ws = wb[sheet_name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= n:
                    break
                rows.append("|".join(str(c) if c is not None else "" for c in row))
            wb.close()
            content = "\n".join(rows)
        except Exception:
            content = str(raw_bytes[:2000])
    else:
        encoding = detect_encoding(raw_bytes)
        text = raw_bytes.decode(encoding, errors="replace")
        lines = text.splitlines()
        content = "\n".join(lines[:n])
    return hashlib.sha256(content.encode()).hexdigest()


def load_raw_head(raw_bytes: bytes, filename: str, n: int = 10) -> "pd.DataFrame":
    """Load first n rows of a file with NO skip_rows and NO preprocessing.
    Used for the schema review UI to show the user the raw file structure."""
    import io as _io
    name_lower = filename.lower()
    if name_lower.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(_io.BytesIO(raw_bytes), read_only=True, data_only=True)
            sheet_name = detect_best_sheet(wb)
            wb.close()
        except Exception:
            sheet_name = 0
        df = pd.read_excel(_io.BytesIO(raw_bytes), sheet_name=sheet_name, header=None, nrows=n)
    else:
        encoding = detect_encoding(raw_bytes)
        text = raw_bytes.decode(encoding, errors="replace")
        delimiter = detect_delimiter(text)
        df = pd.read_csv(
            _io.StringIO(text),
            sep=delimiter,
            header=None,
            nrows=n,
            engine="python",
            on_bad_lines="skip",
        )
    # Defensive: coerce column names to str (Excel datetime headers, etc.)
    df.columns = [str(c) for c in df.columns]
    return df


# ── Phase-0 preprocessing ──────────────────────────────────────────────────────

def detect_and_strip_preheader_rows(
    df: pd.DataFrame,
    source_name: str = "",
) -> tuple[pd.DataFrame, int]:
    """Remove spurious metadata rows that appear *before* the actual column header.

    NOTE: This function is currently bypassed when the pre-load density-based
    header detection (detect_skip_rows) returns certain=True. It is kept as a
    fallback for edge cases where pre-load detection fails. The pre-load approach
    is preferred because it avoids loading garbage data into pandas and then
    trying to clean it up post-hoc.

    Banks occasionally export files where the first few rows contain account
    info, report titles, or empty lines before the real transaction table
    header.  pandas reads those rows as data, producing wrong column names and
    extra NaN-heavy rows at the top.

    Algorithm (language-agnostic, statistical):
    1. Reconstruct the pandas-consumed header row as row 0 of a working copy
       (``Unnamed: N`` synthetic column names are treated as empty cells).
    2. Count non-null cells per row → density array.
    3. Compute the median density across all rows.
    4. Walk from row 0 downward; collect contiguous rows whose density is
       below ``median × _PREHEADER_DENSITY_THRESHOLD``.
    5. Safety caps: if the candidate count exceeds ``_PREHEADER_MAX_ROWS``
       **or** ``_PREHEADER_MAX_RATIO × total_rows``, raise ``ValueError``
       (import stops with a descriptive message rather than silently
       producing garbage data).
    6. If ``n_sparse > 0``, reassign column names from the first non-sparse
       row (row index ``n_sparse`` in the working copy) and return the data
       rows that follow.

    Args:
        df: Raw DataFrame as returned by ``pd.read_csv`` / ``pd.read_excel``.
        source_name: Filename used only for log messages.

    Returns:
        ``(trimmed_df, n_skipped)`` where ``n_skipped`` is the number of
        pre-header rows removed.  If nothing was stripped, ``n_skipped == 0``
        and the original DataFrame is returned unchanged.

    Raises:
        ValueError: If the number of sparse rows at the start exceeds the
            safety caps, indicating the file probably does not start with a
            transaction table at all.
    """
    logger.info("── detect_and_strip_preheader_rows [%s] ──", source_name)
    if len(df) < 4:
        logger.info("  DataFrame too short (%d rows), skipping strip.", len(df))
        return df, 0

    total_rows = len(df) + 1  # +1 for the reconstructed header row
    logger.info("  total_rows (incl header)=%d, columns=%s", total_rows, list(df.columns)[:8])

    # Step 1 – Reconstruct the pandas-consumed header row.
    header_values: list = [
        None if re.match(r"^Unnamed: \d+$", str(c)) else str(c)
        for c in df.columns
    ]
    n_cols = len(header_values)
    logger.info("  pandas header (row 0): %s", [str(v)[:25] for v in header_values])

    header_series = pd.Series(header_values, index=df.columns)
    df_full = pd.concat(
        [header_series.to_frame().T, df.reset_index(drop=True)],
        ignore_index=True,
    )

    # Step 2 – Non-null density per row.
    densities: list[float] = [
        row.notna().sum() / n_cols
        for _, row in df_full.iterrows()
    ]

    # Log first 25 rows with density
    for i, d in enumerate(densities[:25]):
        row_vals = [str(v)[:20] for v in df_full.iloc[i] if pd.notna(v)][:5]
        logger.info("  row %2d | density=%.2f | values=%s", i, d, row_vals)

    # Step 3 – Median density.
    med = median(densities)
    threshold = med * _PREHEADER_DENSITY_THRESHOLD
    logger.info("  median_density=%.2f, threshold=%.2f (median × %.2f)",
                med, threshold, _PREHEADER_DENSITY_THRESHOLD)

    # Step 4 – Contiguous sparse rows at the start.
    n_sparse = 0
    for d in densities:
        if d < threshold:
            n_sparse += 1
        else:
            break  # stop at first non-sparse row

    logger.info("  contiguous sparse rows from top: %d", n_sparse)

    if n_sparse == 0:
        logger.info("  → No pre-header rows to strip.")
        return df, 0

    # Step 5 – Safety caps.
    ratio = n_sparse / total_rows
    logger.info("  n_sparse=%d, ratio=%.2f%%, cap_abs=%d, cap_ratio=%.0f%%",
                n_sparse, ratio * 100, _PREHEADER_MAX_ROWS, _PREHEADER_MAX_RATIO * 100)

    if n_sparse > _PREHEADER_MAX_ROWS:
        raise ValueError(
            f"[{source_name}] detect_and_strip_preheader_rows: "
            f"{n_sparse} sparse rows detected at the start, exceeding the "
            f"absolute cap of {_PREHEADER_MAX_ROWS}. "
            "The file may not contain a standard transaction table."
        )
    if ratio > _PREHEADER_MAX_RATIO:
        raise ValueError(
            f"[{source_name}] detect_and_strip_preheader_rows: "
            f"{n_sparse} sparse rows represent {ratio:.1%} of the file, "
            f"exceeding the {_PREHEADER_MAX_RATIO:.0%} safety cap. "
            "The file may not contain a standard transaction table."
        )

    # Step 6 – Reassign column names from the first non-sparse row.
    new_header_row = df_full.iloc[n_sparse]
    new_columns = [
        (str(v).strip() if pd.notna(v) and str(v).strip() else f"Unnamed: {i}")
        for i, v in enumerate(new_header_row)
    ]
    logger.info("  → New header from row %d: %s", n_sparse, new_columns[:8])

    result = df_full.iloc[n_sparse + 1:].copy()
    result.columns = new_columns
    result = result.reset_index(drop=True)

    logger.info(
        "[%s] Stripped %d pre-header row(s) (density threshold=%.2f × median %.2f). "
        "Result: %d rows, columns=%s",
        source_name, n_sparse, _PREHEADER_DENSITY_THRESHOLD, med,
        len(result), list(result.columns)[:8],
    )
    return result, n_sparse


# ── Footer detection constants ────────────────────────────────────────────────
_FOOTER_MAX_ROWS: int = 5
_FOOTER_MAX_RATIO: float = 0.05  # 5 % of total rows


def detect_and_strip_footer_rows(
    df: pd.DataFrame,
    source_name: str = "",
) -> tuple[pd.DataFrame, int]:
    """Remove summary/totale rows at the bottom of the table using IQR outlier detection.

    Algorithm (language-agnostic, statistical):
    1. Compute per-row density (non-null fraction).
    2. Compute Q1, Q3, IQR = Q3 - Q1.
    3. lower_fence = Q1 - 1.5 × IQR.
    4. Scan from the bottom: collect contiguous rows whose density < lower_fence.
    5. Safety caps: max _FOOTER_MAX_ROWS or _FOOTER_MAX_RATIO × total_rows.

    Returns:
        (trimmed_df, n_stripped) — n_stripped == 0 if nothing was removed.
    """
    if len(df) < 5:
        return df, 0

    n_cols = len(df.columns)
    if n_cols == 0:
        return df, 0

    densities = [row.notna().sum() / n_cols for _, row in df.iterrows()]

    sorted_d = sorted(densities)
    n = len(sorted_d)
    q1 = sorted_d[n // 4]
    q3 = sorted_d[(3 * n) // 4]
    iqr = q3 - q1
    lower_fence = q1 - 1.5 * iqr

    # Scan from bottom upward: collect contiguous sparse rows
    n_footer = 0
    for d in reversed(densities):
        if d < lower_fence:
            n_footer += 1
        else:
            break

    if n_footer == 0:
        return df, 0

    # Safety caps
    max_allowed = min(_FOOTER_MAX_ROWS, int(len(df) * _FOOTER_MAX_RATIO))
    n_footer = min(n_footer, max(max_allowed, 1))

    result = df.iloc[:-n_footer].copy().reset_index(drop=True)
    logger.info(
        "[%s] Stripped %d footer row(s) (IQR lower_fence=%.2f, Q1=%.2f, Q3=%.2f)",
        source_name, n_footer, lower_fence, q1, q3,
    )
    return result, n_footer


def drop_low_variability_columns(
    df: pd.DataFrame,
    source_name: str = "",
    min_ratio: float = _LOW_VARIABILITY_RATIO,
) -> tuple[pd.DataFrame, list[str]]:
    """Remove columns whose value diversity is too low to carry transaction data.

    Columns like "Nome titolare" (always the same name) or "Numero carta"
    (same masked PAN on every row) add noise without information.  A column
    is considered metadata/constant if::

        nunique(col) / nrows < min_ratio

    At least 2 columns are always preserved so downstream processing has
    something to work with.

    Args:
        df: DataFrame (may already be pre-header-stripped).
        source_name: Filename used only for log messages.
        min_ratio: Fraction threshold (default ``_LOW_VARIABILITY_RATIO``
            = 1.5 %).

    Returns:
        ``(cleaned_df, dropped_column_names)``.  If nothing was dropped,
        ``dropped_column_names`` is an empty list.
    """
    if len(df) < 2 or len(df.columns) <= 2:
        return df, []

    nrows = len(df)
    to_drop: list[str] = []

    for col in df.columns:
        ratio = df[col].nunique(dropna=True) / nrows
        if ratio < min_ratio:
            to_drop.append(col)

    # Never drop below 2 columns.
    max_droppable = len(df.columns) - 2
    if len(to_drop) > max_droppable:
        to_drop = to_drop[:max_droppable]

    if not to_drop:
        return df, []

    result = df.drop(columns=to_drop)
    logger.info(
        "[%s] Dropped %d low-variability column(s): %s",
        source_name, len(to_drop), to_drop,
    )
    return result, to_drop


# ── Internal transfer detection (RF-04) ──────────────────────────────────────

def _build_owner_name_regex(owner_names: list[str]) -> re.Pattern | None:
    """Build a regex that matches any token-permutation of each owner name.

    For "Mario Rossi" generates both "Mario Rossi" and "Rossi Mario" patterns
    so the description matches regardless of the order banks write the name.
    Each token permutation is joined with \\s+ to tolerate extra spaces.
    Single-token names (e.g. a company abbreviation) are matched as-is.
    """
    if not owner_names:
        return None
    patterns: list[str] = []
    for name in owner_names:
        tokens = name.strip().split()
        if not tokens:
            continue
        if len(tokens) == 1:
            patterns.append(re.escape(tokens[0]))
        else:
            for perm in _permutations(tokens):
                patterns.append(r"\s+".join(re.escape(t) for t in perm))
    if not patterns:
        return None
    return re.compile("|".join(f"(?:{p})" for p in patterns), re.IGNORECASE)


def detect_internal_transfers(
    df: pd.DataFrame,
    epsilon: Decimal = Decimal("0.01"),
    delta_days: int = 5,
    epsilon_strict: Decimal = Decimal("0.005"),
    delta_days_strict: int = 1,
    keyword_patterns: list[str] | None = None,
    require_keyword_confirmation: bool = True,
    owner_names: list[str] | None = None,
) -> pd.DataFrame:
    """
    Mark pairs of transactions as internal transfers (RF-04).

    Adds columns:
      - transfer_pair_id: shared ID for the matched pair (or None)
      - tx_type: updated to internal_out / internal_in where matched
      - transfer_confidence: Confidence enum value

    The df must have columns: id, date, amount (Decimal), description,
    account_label, tx_type.
    """
    df = df.copy()
    if "transfer_pair_id" not in df.columns:
        df["transfer_pair_id"] = None
    if "transfer_confidence" not in df.columns:
        df["transfer_confidence"] = None

    keyword_re = None
    if keyword_patterns:
        keyword_re = re.compile('|'.join(re.escape(p) for p in keyword_patterns), re.IGNORECASE)

    def _keyword_match(text: str) -> bool:
        if not keyword_re:
            return False
        return bool(keyword_re.search(text or ""))

    def _high_sym(ai: Decimal, di: date, aj: Decimal, dj: date) -> bool:
        return (
            abs(ai + aj) <= epsilon_strict
            and abs((di - dj).days) <= delta_days_strict
        )

    indices = df.index.tolist()
    already_paired: set = set()

    for i, j in combinations(indices, 2):
        if i in already_paired or j in already_paired:
            continue
        ri, rj = df.loc[i], df.loc[j]
        # Skip same account
        if ri.get("account_label") == rj.get("account_label"):
            continue

        ai: Decimal = ri["amount"] if isinstance(ri["amount"], Decimal) else Decimal(str(ri["amount"] or 0))
        aj: Decimal = rj["amount"] if isinstance(rj["amount"], Decimal) else Decimal(str(rj["amount"] or 0))
        di: date = ri["date"] if isinstance(ri["date"], date) else pd.to_datetime(ri["date"]).date()
        dj: date = rj["date"] if isinstance(rj["date"], date) else pd.to_datetime(rj["date"]).date()

        amount_match = abs(ai + aj) <= epsilon
        date_match = abs((di - dj).days) <= delta_days

        if not (amount_match and date_match):
            continue

        kw_i = _keyword_match(str(ri.get("description", "")))
        kw_j = _keyword_match(str(rj.get("description", "")))
        high_sym = _high_sym(ai, di, aj, dj)

        if kw_i or kw_j:
            confidence = Confidence.high
        elif high_sym:
            confidence = Confidence.medium
        else:
            continue

        if require_keyword_confirmation and confidence == Confidence.medium:
            # Mark for review but don't assign tx_type yet
            pair_id = f"transfer_{i}_{j}"
            df.at[i, "transfer_pair_id"] = pair_id
            df.at[j, "transfer_pair_id"] = pair_id
            df.at[i, "transfer_confidence"] = confidence.value
            df.at[j, "transfer_confidence"] = confidence.value
            already_paired.update([i, j])
            continue

        pair_id = f"transfer_{i}_{j}"
        out_idx = i if ai < 0 else j
        in_idx = j if ai < 0 else i

        df.at[out_idx, "tx_type"] = TransactionType.internal_out.value
        df.at[in_idx, "tx_type"] = TransactionType.internal_in.value
        df.at[i, "transfer_pair_id"] = pair_id
        df.at[j, "transfer_pair_id"] = pair_id
        df.at[i, "transfer_confidence"] = confidence.value
        df.at[j, "transfer_confidence"] = confidence.value
        already_paired.update([i, j])

    # Owner-name pass: any unpaired transaction whose description contains an
    # owner name is marked directly as internal_out / internal_in (no pairing
    # needed — the owner is the other side of the transfer).
    # Owner-name pass: match any token-permutation of each owner name so that
    # both "Mario Rossi" and "ROSSI MARIO" (common in bank exports) are caught.
    owner_re = _build_owner_name_regex(owner_names) if owner_names else None
    if owner_re:
        for idx in df.index:
            if idx in already_paired:
                continue
            desc = str(df.at[idx, "description"] or "")
            if owner_re.search(desc):
                amt = df.at[idx, "amount"]
                if not isinstance(amt, Decimal):
                    amt = Decimal(str(amt or 0))
                df.at[idx, "tx_type"] = (
                    TransactionType.internal_out.value if amt < 0
                    else TransactionType.internal_in.value
                )
                df.at[idx, "transfer_confidence"] = Confidence.high.value
                logger.info(
                    f"detect_internal_transfers: owner-name match → "
                    f"idx={idx} desc='{desc[:40]}' "
                    f"type={df.at[idx, 'tx_type']}"
                )

    return df


# ── Card settlement reconciliation (RF-03) ────────────────────────────────────

def find_card_settlement_matches(
    settlements: list[dict],
    card_transactions: list[dict],
    epsilon: Decimal = Decimal("0.01"),
    window_days: int = 45,
    max_gap_days: int = 5,
    boundary_k: int = 10,
) -> list[dict]:
    """
    Match card_settlement rows from the bank account with card_tx rows.

    Returns a list of reconciliation records:
      {settlement_id, matched_ids: list[str], delta: Decimal, method: str}

    Three-phase algorithm:
      1. Temporal window filter
      2. Sliding window contiguous matching
      3. Subset sum at boundary
    """
    results = []

    for settlement in settlements:
        s_id = settlement["id"]
        s_amount = abs(settlement["amount"])
        s_date = settlement["date"] if isinstance(settlement["date"], date) else pd.to_datetime(settlement["date"]).date()

        # Phase 1: temporal window
        candidates = [
            tx for tx in card_transactions
            if not tx.get("reconciled", False)
        ]
        window_start = s_date - timedelta(days=window_days)
        window_end = s_date + timedelta(days=7)
        windowed = [
            tx for tx in candidates
            if window_start <= (tx["date"] if isinstance(tx["date"], date) else pd.to_datetime(tx["date"]).date()) <= window_end
        ]

        # Phase 2: sliding window (contiguous subsets respecting max_gap_days)
        match = _sliding_window_match(windowed, s_amount, epsilon, max_gap_days)
        if match:
            results.append({
                "settlement_id": s_id,
                "matched_ids": [tx["id"] for tx in match],
                "delta": s_amount - sum(abs(tx["amount"]) for tx in match),
                "method": "sliding_window",
            })
            for tx in match:
                tx["reconciled"] = True
            continue

        # Phase 3: subset sum at boundary
        boundary_txs = _get_boundary_transactions(windowed, s_date, boundary_k)
        match = _subset_sum_match(boundary_txs, s_amount, epsilon)
        if match:
            results.append({
                "settlement_id": s_id,
                "matched_ids": [tx["id"] for tx in match],
                "delta": s_amount - sum(abs(tx["amount"]) for tx in match),
                "method": "subset_sum",
            })
            for tx in match:
                tx["reconciled"] = True

    return results


def _sliding_window_match(
    transactions: list[dict],
    target: Decimal,
    epsilon: Decimal,
    max_gap: int,
) -> list[dict] | None:
    """Find a contiguous subset whose sum ≈ target."""
    txs = sorted(
        transactions,
        key=lambda t: t["date"] if isinstance(t["date"], date) else pd.to_datetime(t["date"]).date(),
    )
    n = len(txs)
    for start in range(n):
        total = Decimal(0)
        subset = []
        for end in range(start, n):
            d_cur = txs[end]["date"] if isinstance(txs[end]["date"], date) else pd.to_datetime(txs[end]["date"]).date()
            if subset:
                d_prev = txs[end - 1]["date"] if isinstance(txs[end - 1]["date"], date) else pd.to_datetime(txs[end - 1]["date"]).date()
                if (d_cur - d_prev).days > max_gap:
                    break
            total += abs(txs[end]["amount"]) if isinstance(txs[end]["amount"], Decimal) else Decimal(str(abs(txs[end]["amount"])))
            subset.append(txs[end])
            if abs(total - target) <= epsilon:
                return subset
            if total > target + epsilon:
                break
    return None


def _get_boundary_transactions(transactions: list[dict], ref_date: date, k: int) -> list[dict]:
    sorted_txs = sorted(
        transactions,
        key=lambda t: t["date"] if isinstance(t["date"], date) else pd.to_datetime(t["date"]).date(),
    )
    before = [tx for tx in sorted_txs if (tx["date"] if isinstance(tx["date"], date) else pd.to_datetime(tx["date"]).date()) < ref_date]
    after = [tx for tx in sorted_txs if (tx["date"] if isinstance(tx["date"], date) else pd.to_datetime(tx["date"]).date()) >= ref_date]
    return before[-k:] + after[:k]


def _subset_sum_match(
    transactions: list[dict],
    target: Decimal,
    epsilon: Decimal,
) -> list[dict] | None:
    """Exhaustive subset sum (O(2^n), n ≤ 2k = 20 by default → safe)."""
    n = len(transactions)
    for r in range(1, n + 1):
        for subset in combinations(range(n), r):
            total = sum(
                abs(transactions[i]["amount"]) if isinstance(transactions[i]["amount"], Decimal)
                else Decimal(str(abs(transactions[i]["amount"])))
                for i in subset
            )
            if abs(total - target) <= epsilon:
                return [transactions[i] for i in subset]
    return None


# ── Card within-file balance row removal (Case 5) ────────────────────────────

def remove_card_balance_row(
    transactions: list[dict],
    epsilon: Decimal = Decimal("0.01"),
    owner_name_label: str | None = None,
) -> tuple[list[dict], bool]:
    """Handle the single balance/totale summary row from a card file if present.

    Some card exports include a row whose |amount| equals the sum of |all other
    amounts| (the file total). Including it would double-count every expense.

    Detection rule (requires ≥ 3 transactions):
        For each candidate row i:
            sum_others = Σ |amount_j| for j ≠ i
            if ||amount_i| - sum_others| ≤ epsilon  →  row i is the balance row

    Behaviour:
        - If owner_name_label is provided: rename the row's description to the
          owner name instead of removing it, so the transfer detector can later
          mark it as giroconto (internal transfer).
        - Otherwise: remove the row (legacy behaviour, avoids double-counting).

    Only the FIRST such row is handled (there should be at most one).

    Returns:
        (transactions, was_found)
    """
    if len(transactions) < 3:
        return transactions, False

    amounts = [
        abs(tx["amount"]) if isinstance(tx["amount"], Decimal) else Decimal(str(abs(tx["amount"])))
        for tx in transactions
    ]
    total = sum(amounts)

    for i, tx in enumerate(transactions):
        amt_i = amounts[i]
        sum_others = total - amt_i
        if abs(amt_i - sum_others) <= epsilon:
            if owner_name_label:
                logger.info(
                    f"remove_card_balance_row: balance/totale row relabelled as '{owner_name_label}' "
                    f"id={tx.get('id', '?')} amount={amt_i} ≈ sum_others={sum_others} "
                    f"(source={tx.get('source_file', '?')})"
                )
                transactions[i]["description"] = owner_name_label
                return transactions, True
            else:
                logger.info(
                    f"remove_card_balance_row: removed balance/totale row "
                    f"id={tx.get('id', '?')} amount={amt_i} ≈ sum_others={sum_others} "
                    f"(source={tx.get('source_file', '?')})"
                )
                return transactions[:i] + transactions[i + 1:], True

    return transactions, False


# ── 3-Phase Footer Stripping ─────────────────────────────────────────────────
#
# Phase 1: Structural filter (column density, schema-aware)
# Phase 2: Semantic extraction with LLM (last 10 rows)
# Phase 3: Reuse stored textual patterns
# ---------------------------------------------------------------------------

# Regex for stripping dates and numbers from description text to build patterns
_DATE_PATTERN = re.compile(r"\d{1,4}[/\-\.]\d{1,2}[/\-\.]\d{1,4}")
_NUMBER_PATTERN = re.compile(r"[\d.,]+")
_MULTI_SPACE = re.compile(r"\s+")

# Multilingual keywords that signal a suspicious (possible footer) row
_FOOTER_SUSPECT_KEYWORDS = frozenset({
    "totale", "totali", "total", "totaux",
    "saldo", "saldi", "balance", "solde",
    "firma", "firme", "signature", "unterschrift",
    "timbro", "stamp", "cachet",
    "pagina", "page", "seite",
    "estratto", "extract", "extrait",
    "riepilogo", "summary", "résumé", "zusammenfassung",
})

_FOOTER_MIN_PATTERN_LEN = 3  # patterns shorter than this are discarded


def _resolve_description_col(schema) -> Optional[str]:
    """Return the primary description column name from a DocumentSchema."""
    desc_cols = getattr(schema, "description_cols", None)
    if desc_cols:
        return desc_cols[0]
    return getattr(schema, "description_col", None)


def _normalize_description_to_pattern(text: str) -> str:
    """Strip dates, numbers, and extra whitespace from description text → residual pattern."""
    t = _DATE_PATTERN.sub("", text)
    t = _NUMBER_PATTERN.sub("", t)
    t = _MULTI_SPACE.sub(" ", t).strip().lower()
    return t


def strip_footer_phase1(
    df: pd.DataFrame,
    schema,
    source_name: str = "",
) -> tuple[pd.DataFrame, int]:
    """Phase 1: Structural filter — remove bottom rows with NA in mandatory columns.

    For 3-column schemas (amount_col set): date, description, amount all required.
    For 4-column schemas (debit_col + credit_col): date and description required.

    Scans from the bottom upward collecting contiguous rows that violate the
    constraint. Applies the same safety caps as the old IQR method.
    """
    if len(df) < 3:
        return df, 0

    desc_col = _resolve_description_col(schema)

    # Determine mandatory columns based on schema shape
    if getattr(schema, "amount_col", None):
        # 3-column schema: date + description + amount all mandatory
        mandatory = [schema.date_col]
        if desc_col and desc_col in df.columns:
            mandatory.append(desc_col)
        if schema.amount_col in df.columns:
            mandatory.append(schema.amount_col)
    elif getattr(schema, "debit_col", None) and getattr(schema, "credit_col", None):
        # 4-column schema: date + description mandatory
        mandatory = [schema.date_col]
        if desc_col and desc_col in df.columns:
            mandatory.append(desc_col)
    else:
        # Unknown shape — fall back to date-only check
        mandatory = [schema.date_col]

    # Filter to columns that actually exist in the DataFrame
    mandatory = [c for c in mandatory if c in df.columns]
    if not mandatory:
        return df, 0

    # Scan from bottom: count contiguous rows where any mandatory column is NA
    n_footer = 0
    for i in range(len(df) - 1, -1, -1):
        row = df.iloc[i]
        has_na = any(pd.isna(row.get(c)) for c in mandatory)
        if has_na:
            n_footer += 1
        else:
            break

    if n_footer == 0:
        return df, 0

    # Safety caps
    max_allowed = min(_FOOTER_MAX_ROWS, int(len(df) * _FOOTER_MAX_RATIO))
    n_footer = min(n_footer, max(max_allowed, 1))

    result = df.iloc[:-n_footer].copy().reset_index(drop=True)
    logger.info(
        "[%s] Phase 1 footer strip: removed %d row(s) with NA in mandatory columns %s",
        source_name, n_footer, mandatory,
    )
    return result, n_footer


def strip_footer_phase2_llm(
    df: pd.DataFrame,
    schema,
    llm_backend,
    sanitize_config=None,
    source_name: str = "",
) -> tuple[pd.DataFrame, list[str], int]:
    """Phase 2: Semantic extraction — ask LLM to identify footer rows among the last 10.

    Returns:
        (trimmed_df, new_patterns, n_stripped)
        new_patterns: list of normalised text patterns extracted from identified footers.
    """
    import json as _json
    from pathlib import Path

    if len(df) < 1:
        return df, [], 0

    tail_size = min(10, len(df))
    tail_start = len(df) - tail_size
    tail_df = df.iloc[tail_start:]

    desc_col = _resolve_description_col(schema)

    # Build row representations for the LLM
    rows_for_llm = []
    for local_idx, (_, row) in enumerate(tail_df.iterrows()):
        row_dict = {}
        for col in df.columns:
            val = row.get(col)
            row_dict[col] = "" if pd.isna(val) else str(val)[:200]
        # Sanitize descriptions for remote backends
        if sanitize_config and desc_col and desc_col in row_dict:
            from core.sanitizer import redact_pii
            row_dict[desc_col] = redact_pii(row_dict[desc_col], sanitize_config)
        rows_for_llm.append({"index": local_idx, **row_dict})

    # Load prompt
    _prompts_file = Path(__file__).parent.parent / "prompts" / "footer_detector.json"
    with open(_prompts_file, encoding="utf-8") as f:
        prompts = _json.load(f)

    system_prompt = prompts["system"]

    # Build amount info string
    if getattr(schema, "amount_col", None):
        amount_info = f"single column: {schema.amount_col}"
    elif getattr(schema, "debit_col", None):
        amount_info = f"debit: {schema.debit_col}, credit: {getattr(schema, 'credit_col', '?')}"
    else:
        amount_info = "unknown"

    user_prompt = prompts["user_template"].format(
        date_col=schema.date_col,
        description_col=desc_col or "N/A",
        amount_info=amount_info,
        rows_json=_json.dumps(rows_for_llm, ensure_ascii=False, indent=2),
    )

    response_schema = prompts["response_schema"]

    try:
        result = llm_backend.complete_structured(system_prompt, user_prompt, response_schema)
    except Exception as exc:
        logger.warning("[%s] Phase 2 LLM footer detection failed: %s", source_name, exc)
        raise

    footer_indices_local = []
    for item in result.get("footer_rows", []):
        idx = item.get("index")
        if isinstance(idx, int) and 0 <= idx < tail_size:
            footer_indices_local.append(idx)

    if not footer_indices_local:
        return df, [], 0

    # Extract patterns from footer rows
    new_patterns = []
    for local_idx in footer_indices_local:
        row = tail_df.iloc[local_idx]
        desc = str(row.get(desc_col, "") or "") if desc_col else ""
        pattern = _normalize_description_to_pattern(desc)
        if len(pattern) >= _FOOTER_MIN_PATTERN_LEN:
            new_patterns.append(pattern)

    # Remove footer rows from the full DataFrame
    global_indices = [tail_start + i for i in footer_indices_local]
    result_df = df.drop(df.index[global_indices]).reset_index(drop=True)

    n_stripped = len(footer_indices_local)
    logger.info(
        "[%s] Phase 2 LLM footer strip: removed %d row(s), learned %d pattern(s)",
        source_name, n_stripped, len(new_patterns),
    )
    return result_df, new_patterns, n_stripped


def strip_footer_phase3_patterns(
    df: pd.DataFrame,
    schema,
    stored_patterns: list[str],
    source_name: str = "",
) -> tuple[pd.DataFrame, list[int], int]:
    """Phase 3: Reuse stored patterns — match against the last 10 rows.

    Returns:
        (trimmed_df, unmatched_suspect_indices, n_stripped)
        unmatched_suspect_indices: global DF indices of rows in the tail that
        look suspicious (contain footer keywords) but were NOT matched by any
        stored pattern — used to decide whether to trigger Phase 2.
    """
    if len(df) < 1 or not stored_patterns:
        return df, [], 0

    tail_size = min(10, len(df))
    tail_start = len(df) - tail_size

    desc_col = _resolve_description_col(schema)

    matched_global = []
    unmatched_suspect = []

    for i in range(tail_size):
        global_idx = tail_start + i
        row = df.iloc[global_idx]
        desc = str(row.get(desc_col, "") or "") if desc_col else ""
        normalised = _normalize_description_to_pattern(desc)

        # Check if any stored pattern matches
        pattern_hit = False
        for pat in stored_patterns:
            if pat and pat in normalised:
                pattern_hit = True
                break

        if pattern_hit:
            matched_global.append(global_idx)
        else:
            # Check if this row looks suspicious based on keywords
            desc_lower = desc.lower()
            if any(kw in desc_lower for kw in _FOOTER_SUSPECT_KEYWORDS):
                unmatched_suspect.append(global_idx)

    if not matched_global:
        return df, unmatched_suspect, 0

    result_df = df.drop(df.index[matched_global]).reset_index(drop=True)
    n_stripped = len(matched_global)
    logger.info(
        "[%s] Phase 3 pattern footer strip: removed %d row(s), %d unmatched suspect(s)",
        source_name, n_stripped, len(unmatched_suspect),
    )
    return result_df, unmatched_suspect, n_stripped
