#!/usr/bin/env python3
"""Generate synthetic Italian movements files for Spendify import testing.

Produces ~50 CSV/XLSX files simulating 9 Italian financial instruments
across 3 sizes (S/M/L) plus format variants. Each file has realistic
Italian banking transaction descriptions, preheader rows, and footers.

Usage:
    python tests/generate_synthetic_files.py
"""
from __future__ import annotations

import csv
import random
import sys
from dataclasses import dataclass, field
from datetime import date, timedelta, datetime
from decimal import Decimal
from pathlib import Path
from typing import Literal

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Border, Side, Font, PatternFill

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed – XLSX files will be skipped.")

# ── Deterministic seed ────────────────────────────────────────────────────
random.seed(42)

# ── Output directory ──────────────────────────────────────────────────────
OUTPUT_DIR = Path(__file__).resolve().parent / "generated_files"

# ── Date range ────────────────────────────────────────────────────────────
DATE_START = date(2026, 1, 1)
DATE_END = date(2026, 2, 28)

# ── Financial instruments ─────────────────────────────────────────────────

@dataclass
class Instrument:
    id: str
    tipo: str
    intestatario: str
    banca: str
    doc_type: str
    iban: str
    cc_number: str  # last 7 digits, used in giroconti


INSTRUMENTS = [
    Instrument("CC-1", "Conto corrente", "Marco Rossi + Laura Bianchi",
               "Intesa Sanpaolo", "bank_account",
               "IT60 X054 2811 1010 0000 0123 456", "0123456"),
    Instrument("CC-2", "Conto corrente", "Marco Rossi (P.IVA)",
               "Banca Sella", "bank_account",
               "IT28 W030 6909 6061 0000 0075 849", "0075849"),
    Instrument("CC-3", "Conto corrente", "Rosa Verdi",
               "Banco BPM", "bank_account",
               "IT40 S050 3401 6000 0000 0036 218", "0036218"),
    Instrument("CRED-1", "Carta di credito", "Marco Rossi",
               "American Express", "credit_card",
               "", "3742-1234"),
    Instrument("CRED-2", "Carta di credito", "Laura Bianchi",
               "CartaSì/Nexi", "credit_card",
               "", "5412-7890"),
    Instrument("RIC-1", "Carta ricaricabile", "Sofia Rossi",
               "PostePay Evolution", "prepaid_card",
               "IT76 P076 0101 6000 0109 8765 432", "9876543"),
    Instrument("RIC-2", "Carta ricaricabile", "Marco Rossi",
               "Hype", "prepaid_card",
               "IT86 I069 5005 8000 0005 5123 456", "5123456"),
    Instrument("RIC-3", "Carta ricaricabile", "Laura Bianchi",
               "Revolut", "prepaid_card",
               "LT12 3456 7890 1234 5678", "3456789"),
    Instrument("RISP-1", "Conto risparmio", "Marco Rossi + Laura Bianchi",
               "Conto Arancio", "savings_account",
               "IT33 D034 4001 6000 0000 0468 012", "0468012"),
]

INSTRUMENT_MAP = {i.id: i for i in INSTRUMENTS}

# ── Column name pools ─────────────────────────────────────────────────────

DATE_COL_NAMES = [
    "Data operazione", "Data contabile", "Data registrazione",
    "Data mov.", "Data",
]
DATE_VALUTA_NAMES = [
    "Data valuta", "Data val.",
]
DESCRIPTION_COL_NAMES = [
    "Descrizione", "Causale", "Tipologia",
    "Descrizione operazione", "Movimento", "Dettaglio",
]
CURRENCY_COL_NAMES = [
    "Divisa", "Valuta", None,  # None = absent
]

# Amount format definitions
# "signed_single": one column, negative = expense
# "debit_credit_split": two columns (positive values)
# "debit_credit_signed": two columns, debit has negative sign
# "positive_only": one column, all positive (credit cards)

AMOUNT_SINGLE_NAMES = ["Importo", "Movimento", "Dare/Avere"]
AMOUNT_DEBIT_NAMES = ["Uscite", "Dare", "Addebiti"]
AMOUNT_CREDIT_NAMES = ["Entrate", "Avere", "Accrediti"]

# ── Transaction description templates ─────────────────────────────────────

# --- EXPENSES by category ---

EXPENSE_GROCERY = [
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) ESSELUNGA MILANO",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) CONAD CITY ROMA",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) LIDL ITALIA SRL",
    "PAGAMENTO CONTACTLESS CARTA **** {card4} ESSELUNGA S.P.A.",
    "PAGAMENTO CONTACTLESS CARTA **** {card4} CONAD SUPERSTORE",
    "PAGAMENTO CONTACTLESS CARTA **** {card4} LIDL ITALIA",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) MERCATO RIONALE PORTA PALAZZO",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) MACELLERIA BONI",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) PANIFICIO TOSI",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) PESCHERIA DEL PORTO",
]

EXPENSE_RESTAURANT = [
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) RISTORANTE DA GINO",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) PIZZERIA NAPOLI",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) BAR CENTRALE",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) MCDONALD'S ITALIA",
    "Addebito SDD - DELIVEROO ITALY SRL",
    "Addebito SDD - JUST EAT ITALY SRL",
    "PAGAMENTO CONTACTLESS CARTA **** {card4} TRATTORIA LA PERGOLA",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) ROSTICCERIA DA MARIO",
]

EXPENSE_HOUSING = [
    "Addebito SDD - BANCA INTESA SANPAOLO MUTUO RAT. {ref}",
    "Disposizione - RIF:{ref} BEN. CONDOMINIO VIA ROMA 15",
    "Disposizione - RIF:{ref} BEN. CONDOMINIO RESIDENZA MARE",
    "Addebito SDD - ENEL SERVIZIO ELETTRICO NAZ.",
    "Addebito SDD - ENI PLENITUDE GAS E LUCE",
    "Addebito SDD - HERA COMM S.R.L. ACQUA",
    "Addebito SDD - GRUPPO HERA TARI 2026",
    "Disposizione - RIF:{ref} BEN. IDRAULICO ROSSI PAOLO",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) IKEA ITALIA RETAIL",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) LEROY MERLIN ITALIA",
]

EXPENSE_TRANSPORT = [
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) DISTRIBUTORE ENI",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) STAZIONE Q8",
    "Addebito SDD - TELEPASS S.P.A. PEDAGGI",
    "Addebito SDD - SARA ASSICURAZIONI RCA AUTO",
    "Disposizione - RIF:{ref} BEN. OFFICINA MECCANICA BRAMBILLA",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) TRENITALIA S.P.A.",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) ITALO NTV S.P.A.",
    "Addebito SDD - ACI BOLLO AUTO TARGA {plate}",
]

EXPENSE_HEALTH = [
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) FARMACIA DR ROSSI",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) FARMACIA COMUNALE",
    "Disposizione - RIF:{ref} BEN. DR VERDI ODONTOIATRA",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) CENTRO DIAGNOSTICO SAN DONATO",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) OTTICA AVANZI",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) CENTRO MEDICO SANTAGOSTINO",
]

EXPENSE_EDUCATION = [
    "Addebito SDD - MENSA SCOLASTICA COMUNE DI MILANO",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) LIBRERIA FELTRINELLI",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) CARTOLIBRERIA IL PAPIRO",
    "Disposizione - RIF:{ref} BEN. UNIVERSITA' DEGLI STUDI DI BOLOGNA",
    "Disposizione - RIF:{ref} BEN. AFFITTO APPARTAMENTO SOFIA BOLOGNA",
]

EXPENSE_CLOTHING = [
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) ZARA ITALIA SRL",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) H&M HENNES AND MAURITZ",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) DECATHLON ASSAGO",
    "PAGAMENTO CONTACTLESS CARTA **** {card4} OVS S.P.A.",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) CALZEDONIA GROUP",
]

EXPENSE_TECH = [
    "AMAZON.IT MARKETPLACE - ORDINE 402-{ref}",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) MEDIAWORLD CARUGATE",
    "Addebito SDD - NETFLIX INTERNATIONAL B.V.",
    "Addebito SDD - SPOTIFY AB",
    "Addebito SDD - APPLE.COM/BILL",
    "Addebito SDD - GOOGLE *CLOUD STORAGE",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) UNIEURO S.P.A.",
]

EXPENSE_VACATION = [
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) HOTEL BELLAVISTA RIMINI",
    "Addebito SDD - BOOKING.COM B.V.",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) RYANAIR LTD",
    "Addebito SDD - TELEPASS PEDAGGI A14 AUTOSTRADA",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) RISTORANTE IL FARO CESENATICO",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) BAGNO MARINO N.42 RICCIONE",
]

EXPENSE_WEEKEND = [
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) AGRITURISMO IL CASALE",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) B&B LA CASCINA",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) TERME DI SIRMIONE",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) MUSEO NAZIONALE SCIENZA",
]

EXPENSE_PROFESSIONAL = [
    "AMAZON.IT MARKETPLACE HW - ORDINE 402-{ref}",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) ARUBA S.P.A. HOSTING",
    "Addebito SDD - REGISTER.IT S.P.A. DOMINIO",
    "Disposizione - RIF:{ref} BEN. STUDIO COMM.LE FERRI & ASSOCIATI",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) UDEMY.COM FORMAZIONE",
    "Addebito SDD - GITHUB INC. PRO PLAN",
    "Addebito SDD - JETBRAINS S.R.O. INTELLIJ IDEA",
]

EXPENSE_NONNA = [
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) FARMACIA COMUNALE",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) CONAD SUPERSTORE",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) PARRUCCHIERA SONIA",
    "Disposizione - RIF:{ref} BEN. GIARDINIERE COLOMBO LUIGI",
    "Disposizione - RIF:{ref} BEN. ASSISTENTE FAMILIARE MOLDOVEANU",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) MERCERIA LA BOTTEGA",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) EDICOLA PIAZZA DUOMO",
    "PRELIEVO ATM BANCOMAT FIL. 01234 BANCO BPM",
]

EXPENSE_MISC = [
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) TABACCHERIA N.12",
    "PRELIEVO ATM BANCOMAT FIL. 00456 MILANO",
    "Addebito SDD - WIND TRE S.P.A. MOBILE",
    "Addebito SDD - TELECOM ITALIA S.P.A. FIBRA",
    "Pagam. POS - PAGAMENTO POS {amt} EUR DEL {dt} A (ITA) COOP LOMBARDIA",
    "Addebito SDD - GENERALI ITALIA POLIZZA VITA",
    "Disposizione - RIF:{ref} BEN. TINTORIA RAPIDA CENTRO",
]

# --- INCOMES ---

INCOME_SALARY = [
    "ACCREDITO STIPENDIO MESE DI {month} 2026",
    "Bonif. v/fav. - RIF:{ref} ORD. MINISTERO PUBBLICA ISTRUZIONE STIPENDIO",
]

INCOME_PROFESSIONAL = [
    "Bonif. v/fav. - RIF:{ref} ORD. CLIENTE ALFA S.P.A.",
    "Bonif. v/fav. - RIF:{ref} ORD. BETA CONSULTING SRL",
    "Bonif. v/fav. - RIF:{ref} ORD. GAMMA TECHNOLOGIES SRL",
    "Bonif. v/fav. - RIF:{ref} ORD. DELTA SOLUTIONS SPA",
    "Bonif. v/fav. - RIF:{ref} ORD. EPSILON DIGITAL SRL",
]

INCOME_PENSION = [
    "ACCREDITO PENSIONE INPS MESE DI {month} 2026",
]

INCOME_SAVINGS = [
    "ACCREDITO INTERESSI MATURATI TRIMESTRE Q4 2025",
    "ACCREDITO INTERESSI MATURATI TRIMESTRE Q1 2026",
]

INCOME_MISC = [
    "Bonif. v/fav. - RIF:{ref} ORD. AGENZIA ENTRATE RIMBORSO 730",
    "Bonif. v/fav. - RIF:{ref} ORD. REGIONE LOMBARDIA BONUS",
]

# --- GIROCONTI (internal transfers) ---

GIROCONTO_TEMPLATES = {
    # (from_account, to_account) -> list of templates
    ("CC-1", "RIC-1"): [
        "RICARICA POSTEPAY DA CONTO {from_cc} BEN. SOFIA ROSSI",
        "Disposizione - RIF:{ref} BEN. RICARICA POSTEPAY SOFIA",
    ],
    ("CC-1", "CC-2"): [
        "GIROCONTO DA CC {from_cc} A CC {to_cc}",
        "Disposizione - RIF:{ref} BEN. ROSSI MARCO P.IVA GIROCONTO",
    ],
    ("CC-2", "CC-1"): [
        "GIROCONTO DA CC {from_cc} A CC {to_cc}",
        "Bonif. v/fav. - RIF:{ref} ORD. ROSSI MARCO GIROCONTO",
    ],
    ("CC-3", "CC-1"): [
        "Disposizione - RIF:{ref} BEN. ROSSI MARCO CONTRIBUTO NONNA",
        "GIROCONTO DA CC {from_cc} A CC {to_cc}",
    ],
    ("CC-1", "CC-3"): [
        "Disposizione - RIF:{ref} BEN. VERDI ROSA AIUTO NIPOTE",
    ],
    ("CC-1", "RIC-2"): [
        "RICARICA HYPE DA CONTO {from_cc}",
    ],
    ("CC-1", "RIC-3"): [
        "Disposizione - RIF:{ref} BEN. RICARICA REVOLUT LAURA",
    ],
    ("CC-1", "RISP-1"): [
        "GIROCONTO DA CC {from_cc} A CC {to_cc} ACCANTONAMENTO RISPARMIO",
        "Disposizione - RIF:{ref} BEN. CONTO ARANCIO ACCANTONAMENTO",
    ],
    ("RISP-1", "CC-1"): [
        "GIROCONTO DA CC {from_cc} A CC {to_cc} PRELIEVO RISPARMIO",
    ],
    ("CC-2", "RIC-2"): [
        "RICARICA HYPE DA CONTO {from_cc}",
    ],
}

# ── Amount ranges per category ────────────────────────────────────────────

AMOUNT_RANGES: dict[str, tuple[float, float]] = {
    "grocery": (8.0, 180.0),
    "restaurant": (5.0, 85.0),
    "housing_utility": (35.0, 180.0),
    "housing_mortgage": (780.0, 820.0),
    "housing_condo": (100.0, 250.0),
    "housing_maint": (50.0, 400.0),
    "transport_fuel": (30.0, 90.0),
    "transport_toll": (15.0, 60.0),
    "transport_insurance": (300.0, 600.0),
    "transport_train": (12.0, 65.0),
    "transport_mechanic": (80.0, 500.0),
    "health": (8.0, 250.0),
    "education": (15.0, 800.0),
    "clothing": (15.0, 150.0),
    "tech_purchase": (20.0, 400.0),
    "tech_subscription": (4.99, 14.99),
    "vacation": (30.0, 500.0),
    "weekend": (40.0, 250.0),
    "professional": (20.0, 300.0),
    "nonna_small": (3.0, 50.0),
    "nonna_service": (30.0, 200.0),
    "misc": (5.0, 100.0),
    "atm": (50.0, 250.0),
    # Incomes
    "salary": (2050.0, 2150.0),
    "professional_income": (1500.0, 5500.0),
    "pension": (1150.0, 1250.0),
    "savings_interest": (15.0, 80.0),
    "income_misc": (50.0, 500.0),
    # Giroconti
    "giroconto_small": (50.0, 300.0),
    "giroconto_medium": (300.0, 1000.0),
    "giroconto_savings": (400.0, 600.0),
}

# Italian months
MESI = [
    "GENNAIO", "FEBBRAIO", "MARZO", "APRILE", "MAGGIO", "GIUGNO",
    "LUGLIO", "AGOSTO", "SETTEMBRE", "OTTOBRE", "NOVEMBRE", "DICEMBRE",
]


# ── Helpers ───────────────────────────────────────────────────────────────

def _rand_ref() -> str:
    return str(random.randint(210000000, 219999999))


def _rand_card4() -> str:
    return str(random.randint(1000, 9999))


def _rand_plate() -> str:
    letters = "ABCDEFGHJKLMNPRSTVWXYZ"
    return (
        random.choice(letters) + random.choice(letters)
        + str(random.randint(100, 999))
        + random.choice(letters) + random.choice(letters)
    )


def _rand_date() -> date:
    delta = (DATE_END - DATE_START).days
    return DATE_START + timedelta(days=random.randint(0, delta))


def _format_date_desc(d: date) -> str:
    """Format date as dd.mm.yyyy for inside descriptions."""
    return d.strftime("%d.%m.%Y")


def _format_date_cell(d: date) -> str:
    """Format date as dd/mm/yyyy for CSV cells."""
    return d.strftime("%d/%m/%Y")


def _format_amount_italian(val: float) -> str:
    """Format a float as Italian number: 1.234,56"""
    neg = val < 0
    val = abs(val)
    integer_part = int(val)
    decimal_part = round((val - integer_part) * 100)
    if decimal_part >= 100:
        integer_part += 1
        decimal_part -= 100
    # Thousands separator
    int_str = f"{integer_part:,}".replace(",", ".")
    result = f"{int_str},{decimal_part:02d}"
    if neg:
        result = "-" + result
    return result


def _format_amount_desc(val: float) -> str:
    """Format amount for inside description text (always positive)."""
    return _format_amount_italian(abs(val))


def _rand_amount(category: str) -> float:
    lo, hi = AMOUNT_RANGES[category]
    return round(random.uniform(lo, hi), 2)


def _fill_template(template: str, amount: float, d: date) -> str:
    """Fill placeholders in a transaction description template."""
    return template.format(
        amt=_format_amount_desc(amount),
        dt=_format_date_desc(d),
        card4=_rand_card4(),
        ref=_rand_ref(),
        plate=_rand_plate(),
        month=MESI[d.month - 1],
        from_cc="",
        to_cc="",
    )


# ── Transaction generation per account ────────────────────────────────────

@dataclass
class Transaction:
    date: date
    valuta_date: date | None
    description: str
    amount: float  # positive = income, negative = expense
    is_giroconto: bool = False
    expected_category: str = ""  # ground truth category for testing


# ── Category mapping (internal key → user-facing category/subcategory) ────
CATEGORY_MAP: dict[str, str] = {
    # Expenses — aligned with TAXONOMY_DEFAULTS["it"]["expenses"]
    "grocery": "Alimentari/Spesa supermercato",
    "restaurant": "Ristorazione/Ristorante",
    "housing_utility": "Casa/Energia elettrica",
    "housing_mortgage": "Casa/Mutuo / Affitto",
    "housing_condo": "Casa/Condominio",
    "housing_maint": "Casa/Manutenzione e riparazioni",
    "transport_fuel": "Trasporti/Carburante",
    "transport_toll": "Trasporti/Altro trasporti",
    "transport_insurance": "Trasporti/Assicurazione auto",
    "transport_train": "Trasporti/Trasporto pubblico",
    "transport_mechanic": "Trasporti/Manutenzione auto",
    "health": "Salute/Farmaci",
    "education": "Istruzione/Rette scolastiche",
    "clothing": "Abbigliamento/Abbigliamento adulti",
    "tech_purchase": "Svago e tempo libero/Hobby",
    "tech_subscription": "Svago e tempo libero/Streaming / abbonamenti digitali",
    "vacation": "Svago e tempo libero/Viaggi e vacanze",
    "weekend": "Svago e tempo libero/Cinema / teatro / eventi",
    "professional": "Tasse e tributi/Imposte varie",
    "nonna_small": "Alimentari/Altro alimentari",
    "nonna_service": "Cura personale/Parrucchiere / barbiere",
    "misc": "Altro/Spese non classificate",
    "atm": "Altro/Spese non classificate",
    # Incomes — aligned with TAXONOMY_DEFAULTS["it"]["income"]
    "salary": "Lavoro dipendente/Stipendio",
    "professional_income": "Lavoro autonomo/Fattura / parcella",
    "pension": "Prestazioni sociali/Pensione / rendita",
    "savings_interest": "Rendite finanziarie/Interessi attivi",
    "income_misc": "Trasferimenti e rimborsi/Rimborso generico",
    # Giroconti
    "giroconto_small": "Trasferimenti e rimborsi/Giroconto entrata",
    "giroconto_medium": "Trasferimenti e rimborsi/Giroconto entrata",
    "giroconto_savings": "Finanza e assicurazioni/Investimenti / risparmio",
}


def _generate_expense(templates: list[str], category: str, d: date) -> Transaction:
    amount = _rand_amount(category)
    tmpl = random.choice(templates)
    desc = _fill_template(tmpl, amount, d)
    valuta = d + timedelta(days=random.choice([0, 0, 0, 1, 2]))
    return Transaction(date=d, valuta_date=valuta, description=desc, amount=-amount,
                       expected_category=CATEGORY_MAP.get(category, "Altro/Varie"))


def _generate_income(templates: list[str], category: str, d: date) -> Transaction:
    amount = _rand_amount(category)
    tmpl = random.choice(templates)
    desc = _fill_template(tmpl, amount, d)
    valuta = d + timedelta(days=random.choice([0, 0, 1]))
    return Transaction(date=d, valuta_date=valuta, description=desc, amount=amount,
                       expected_category=CATEGORY_MAP.get(category, "Entrate/Altro"))


def _generate_giroconto_pair(
    from_id: str, to_id: str, d: date
) -> tuple[Transaction, Transaction]:
    """Generate a giroconto: outgoing on from_account, incoming on to_account."""
    key = (from_id, to_id)
    templates = GIROCONTO_TEMPLATES.get(key, ["GIROCONTO DA CC {from_cc} A CC {to_cc}"])
    tmpl = random.choice(templates)
    from_inst = INSTRUMENT_MAP[from_id]
    to_inst = INSTRUMENT_MAP[to_id]

    if "RISPARM" in tmpl.upper() or "ACCANTONAMENTO" in tmpl.upper():
        cat = "giroconto_savings"
    elif "POSTEPAY" in tmpl.upper() or "HYPE" in tmpl.upper() or "REVOLUT" in tmpl.upper():
        cat = "giroconto_small"
    else:
        cat = "giroconto_medium"

    amount = _rand_amount(cat)
    desc_out = tmpl.format(
        from_cc=from_inst.cc_number,
        to_cc=to_inst.cc_number,
        ref=_rand_ref(),
        month=MESI[d.month - 1],
        amt=_format_amount_desc(amount),
        dt=_format_date_desc(d),
        card4="",
        plate="",
    )
    # For the receiving side, reverse the description
    desc_in = desc_out.replace("Disposizione", "Bonif. v/fav.").replace("BEN.", "ORD.")
    if desc_in == desc_out:
        # Generic giroconto: same description works both ways
        desc_in = desc_out

    valuta = d + timedelta(days=random.choice([0, 1]))
    expected_cat = CATEGORY_MAP.get(cat, "Giroconto/Trasferimento")
    tx_out = Transaction(date=d, valuta_date=valuta, description=desc_out,
                         amount=-amount, is_giroconto=True,
                         expected_category=expected_cat)
    tx_in = Transaction(date=d, valuta_date=valuta, description=desc_in,
                        amount=amount, is_giroconto=True,
                        expected_category=expected_cat)
    return tx_out, tx_in


# Distribution of expense categories per account type
ACCOUNT_EXPENSE_MIX: dict[str, list[tuple[list[str], str, float]]] = {
    "CC-1": [
        (EXPENSE_GROCERY, "grocery", 0.20),
        (EXPENSE_RESTAURANT, "restaurant", 0.10),
        (EXPENSE_HOUSING, "housing_utility", 0.12),
        (EXPENSE_TRANSPORT, "transport_fuel", 0.08),
        (EXPENSE_HEALTH, "health", 0.05),
        (EXPENSE_EDUCATION, "education", 0.08),
        (EXPENSE_CLOTHING, "clothing", 0.05),
        (EXPENSE_TECH, "tech_subscription", 0.07),
        (EXPENSE_VACATION, "vacation", 0.03),
        (EXPENSE_WEEKEND, "weekend", 0.03),
        (EXPENSE_MISC, "misc", 0.09),
    ],
    "CC-2": [
        (EXPENSE_PROFESSIONAL, "professional", 0.35),
        (EXPENSE_TECH, "tech_purchase", 0.15),
        (EXPENSE_TRANSPORT, "transport_fuel", 0.10),
        (EXPENSE_RESTAURANT, "restaurant", 0.10),
        (EXPENSE_MISC, "misc", 0.10),
        (EXPENSE_HOUSING, "housing_utility", 0.05),
        (EXPENSE_HEALTH, "health", 0.05),
    ],
    "CC-3": [
        (EXPENSE_NONNA, "nonna_small", 0.35),
        (EXPENSE_NONNA, "nonna_service", 0.15),
        (EXPENSE_HEALTH, "health", 0.20),
        (EXPENSE_GROCERY, "grocery", 0.15),
        (EXPENSE_MISC, "misc", 0.10),
    ],
    "CRED-1": [
        (EXPENSE_TECH, "tech_purchase", 0.20),
        (EXPENSE_RESTAURANT, "restaurant", 0.15),
        (EXPENSE_CLOTHING, "clothing", 0.10),
        (EXPENSE_VACATION, "vacation", 0.10),
        (EXPENSE_GROCERY, "grocery", 0.10),
        (EXPENSE_TRANSPORT, "transport_fuel", 0.08),
        (EXPENSE_PROFESSIONAL, "professional", 0.10),
        (EXPENSE_MISC, "misc", 0.07),
    ],
    "CRED-2": [
        (EXPENSE_GROCERY, "grocery", 0.20),
        (EXPENSE_CLOTHING, "clothing", 0.15),
        (EXPENSE_RESTAURANT, "restaurant", 0.15),
        (EXPENSE_HEALTH, "health", 0.10),
        (EXPENSE_TECH, "tech_subscription", 0.10),
        (EXPENSE_MISC, "misc", 0.10),
        (EXPENSE_EDUCATION, "education", 0.05),
    ],
    "RIC-1": [  # Sofia - studentessa
        (EXPENSE_GROCERY, "grocery", 0.20),
        (EXPENSE_RESTAURANT, "restaurant", 0.25),
        (EXPENSE_CLOTHING, "clothing", 0.10),
        (EXPENSE_TECH, "tech_subscription", 0.10),
        (EXPENSE_EDUCATION, "education", 0.10),
        (EXPENSE_TRANSPORT, "transport_train", 0.10),
        (EXPENSE_MISC, "misc", 0.10),
    ],
    "RIC-2": [  # Marco - Hype small purchases
        (EXPENSE_TECH, "tech_subscription", 0.25),
        (EXPENSE_RESTAURANT, "restaurant", 0.20),
        (EXPENSE_GROCERY, "grocery", 0.15),
        (EXPENSE_MISC, "misc", 0.20),
        (EXPENSE_TRANSPORT, "transport_fuel", 0.10),
    ],
    "RIC-3": [  # Laura - Revolut
        (EXPENSE_GROCERY, "grocery", 0.20),
        (EXPENSE_RESTAURANT, "restaurant", 0.20),
        (EXPENSE_CLOTHING, "clothing", 0.15),
        (EXPENSE_HEALTH, "health", 0.10),
        (EXPENSE_MISC, "misc", 0.15),
        (EXPENSE_TECH, "tech_subscription", 0.10),
    ],
    "RISP-1": [],  # No direct expenses from savings
}

ACCOUNT_INCOME_MIX: dict[str, list[tuple[list[str], str, float]]] = {
    "CC-1": [
        (INCOME_SALARY, "salary", 0.50),
        (INCOME_MISC, "income_misc", 0.10),
    ],
    "CC-2": [
        (INCOME_PROFESSIONAL, "professional_income", 0.80),
        (INCOME_MISC, "income_misc", 0.05),
    ],
    "CC-3": [
        (INCOME_PENSION, "pension", 0.70),
        (INCOME_MISC, "income_misc", 0.10),
    ],
    "CRED-1": [],  # Credit cards have no income (payments show as credits but we skip)
    "CRED-2": [],
    "RIC-1": [],  # Rechargeable cards get money via giroconti
    "RIC-2": [],
    "RIC-3": [],
    "RISP-1": [
        (INCOME_SAVINGS, "savings_interest", 0.30),
    ],
}

# Giroconto pairs per account (from_id, to_id)
ACCOUNT_GIROCONTI: dict[str, list[tuple[str, str]]] = {
    "CC-1": [
        ("CC-1", "RIC-1"), ("CC-1", "CC-2"), ("CC-1", "RIC-2"),
        ("CC-1", "RIC-3"), ("CC-1", "RISP-1"), ("CC-1", "CC-3"),
    ],
    "CC-2": [("CC-2", "CC-1"), ("CC-2", "RIC-2")],
    "CC-3": [("CC-3", "CC-1")],
    "CRED-1": [],
    "CRED-2": [],
    "RIC-1": [],  # receives from CC-1
    "RIC-2": [],  # receives from CC-1, CC-2
    "RIC-3": [],  # receives from CC-1
    "RISP-1": [("RISP-1", "CC-1")],
}


def generate_transactions(
    account_id: str, n_transactions: int
) -> list[Transaction]:
    """Generate n_transactions for a given account."""
    txns: list[Transaction] = []
    inst = INSTRUMENT_MAP[account_id]

    expense_mix = ACCOUNT_EXPENSE_MIX.get(account_id, [])
    income_mix = ACCOUNT_INCOME_MIX.get(account_id, [])
    giroconto_pairs = ACCOUNT_GIROCONTI.get(account_id, [])

    # Decide how many of each type
    if inst.doc_type == "credit_card":
        n_expenses = n_transactions
        n_incomes = 0
        n_giroconti = 0
    elif inst.doc_type == "savings_account":
        n_expenses = 0
        n_incomes = min(4, n_transactions // 5)
        n_giroconti = min(8, n_transactions // 3)
        # Rest are incoming giroconti (we generate them from the other side)
        # For savings, most transactions are giroconti IN
        remaining = n_transactions - n_incomes - n_giroconti
        n_giroconti += remaining
    elif inst.doc_type in ("debit_card", "prepaid_card"):
        # Mostly expenses, some incoming giroconti
        n_expenses = int(n_transactions * 0.85)
        n_incomes = 0
        n_giroconti = n_transactions - n_expenses
    else:
        # Bank accounts: mix
        income_pct = sum(w for _, _, w in income_mix)
        giroconto_pct = 0.10 if giroconto_pairs else 0.0
        n_incomes = max(1, int(n_transactions * income_pct))
        n_giroconti = max(0, int(n_transactions * giroconto_pct))
        n_expenses = n_transactions - n_incomes - n_giroconti

    # Generate expenses
    if expense_mix:
        weights = [w for _, _, w in expense_mix]
        total_w = sum(weights)
        weights = [w / total_w for w in weights]
        for _ in range(n_expenses):
            idx = random.choices(range(len(expense_mix)), weights=weights, k=1)[0]
            templates, category, _ = expense_mix[idx]
            d = _rand_date()
            tx = _generate_expense(templates, category, d)
            txns.append(tx)

    # Generate incomes
    if income_mix:
        weights = [w for _, _, w in income_mix]
        total_w = sum(weights)
        weights = [w / total_w for w in weights]
        for _ in range(n_incomes):
            idx = random.choices(range(len(income_mix)), weights=weights, k=1)[0]
            templates, category, _ = income_mix[idx]
            # Income dates: typically 27th (salary) or 1st (pension) of month
            if category == "salary":
                d = date(2026, random.choice([1, 2]), 27)
            elif category == "pension":
                d = date(2026, random.choice([1, 2]), 1)
            else:
                d = _rand_date()
            tx = _generate_income(templates, category, d)
            txns.append(tx)

    # Generate giroconti
    if giroconto_pairs and n_giroconti > 0:
        for _ in range(n_giroconti):
            pair = random.choice(giroconto_pairs)
            d = _rand_date()
            tx_out, tx_in = _generate_giroconto_pair(pair[0], pair[1], d)
            # For this account, add the outgoing side
            if pair[0] == account_id:
                txns.append(tx_out)
            else:
                txns.append(tx_in)

    # For card / savings incoming giroconti, generate incoming records
    if inst.doc_type in ("debit_card", "prepaid_card") and n_giroconti > 0:
        # Find who sends money to this card
        incoming_pairs = [(f, t) for f, t in GIROCONTO_TEMPLATES.keys()
                          if t == account_id]
        if incoming_pairs:
            for _ in range(n_giroconti):
                pair = random.choice(incoming_pairs)
                d = _rand_date()
                _, tx_in = _generate_giroconto_pair(pair[0], pair[1], d)
                txns.append(tx_in)

    if inst.doc_type == "savings_account":
        # Most savings transactions are incoming giroconti
        incoming_pairs = [(f, t) for f, t in GIROCONTO_TEMPLATES.keys()
                          if t == account_id]
        outgoing_pairs = [(f, t) for f, t in GIROCONTO_TEMPLATES.keys()
                          if f == account_id]
        for i in range(n_giroconti):
            d = _rand_date()
            if i % 4 == 0 and outgoing_pairs:
                pair = random.choice(outgoing_pairs)
                tx_out, _ = _generate_giroconto_pair(pair[0], pair[1], d)
                txns.append(tx_out)
            elif incoming_pairs:
                pair = random.choice(incoming_pairs)
                _, tx_in = _generate_giroconto_pair(pair[0], pair[1], d)
                txns.append(tx_in)

    # Trim or pad to exact count
    if len(txns) > n_transactions:
        txns = txns[:n_transactions]
    while len(txns) < n_transactions:
        d = _rand_date()
        if expense_mix:
            idx = random.randint(0, len(expense_mix) - 1)
            templates, category, _ = expense_mix[idx]
            txns.append(_generate_expense(templates, category, d))
        elif income_mix:
            idx = random.randint(0, len(income_mix) - 1)
            templates, category, _ = income_mix[idx]
            txns.append(_generate_income(templates, category, d))

    # Sort by date
    txns.sort(key=lambda t: t.date)
    return txns


# ── File schema: column names and amount format ──────────────────────────

@dataclass
class FileSchema:
    date_col: str
    valuta_col: str | None
    description_col: str
    currency_col: str | None
    amount_format: Literal[
        "signed_single", "debit_credit_split",
        "debit_credit_signed", "positive_only"
    ]
    amount_col: str | None = None  # for single column
    debit_col: str | None = None   # for split
    credit_col: str | None = None  # for split

    @property
    def column_names(self) -> list[str]:
        cols = [self.date_col]
        if self.valuta_col:
            cols.append(self.valuta_col)
        cols.append(self.description_col)
        if self.amount_format in ("signed_single", "positive_only"):
            cols.append(self.amount_col)
        else:
            cols.append(self.debit_col)
            cols.append(self.credit_col)
        if self.currency_col:
            cols.append(self.currency_col)
        return cols

    @property
    def has_debit_credit_split(self) -> bool:
        return self.amount_format in ("debit_credit_split", "debit_credit_signed")


def _make_schema(index: int, is_credit_card: bool = False) -> FileSchema:
    """Create a unique FileSchema based on index for variety."""
    rng = random.Random(42 + index * 7)

    date_col = DATE_COL_NAMES[index % len(DATE_COL_NAMES)]
    valuta_col = DATE_VALUTA_NAMES[index % len(DATE_VALUTA_NAMES)] if rng.random() < 0.6 else None
    desc_col = DESCRIPTION_COL_NAMES[index % len(DESCRIPTION_COL_NAMES)]
    curr_col = CURRENCY_COL_NAMES[index % len(CURRENCY_COL_NAMES)]

    if is_credit_card:
        amount_format = "positive_only"
    else:
        formats = ["signed_single", "debit_credit_split", "debit_credit_signed"]
        amount_format = formats[index % len(formats)]

    if amount_format in ("signed_single", "positive_only"):
        amount_col = AMOUNT_SINGLE_NAMES[index % len(AMOUNT_SINGLE_NAMES)]
        return FileSchema(
            date_col=date_col, valuta_col=valuta_col,
            description_col=desc_col, currency_col=curr_col,
            amount_format=amount_format, amount_col=amount_col,
        )
    else:
        idx_d = index % len(AMOUNT_DEBIT_NAMES)
        return FileSchema(
            date_col=date_col, valuta_col=valuta_col,
            description_col=desc_col, currency_col=curr_col,
            amount_format=amount_format,
            debit_col=AMOUNT_DEBIT_NAMES[idx_d],
            credit_col=AMOUNT_CREDIT_NAMES[idx_d],
        )


# ── Preheader / footer generation ─────────────────────────────────────────

def _generate_preheader(inst: Instrument, n_rows: int) -> list[list[str]]:
    """Generate preheader rows for a file."""
    if n_rows == 0:
        return []

    pool = [
        [f"Intestatario: {inst.intestatario}"],
        [f"Banca: {inst.banca}"],
        [f"Tipo conto: {inst.tipo}"],
        [],  # empty row
        [f"Periodo: 01/01/2026 - 28/02/2026"],
    ]
    if inst.iban:
        pool.insert(2, [f"IBAN: {inst.iban}"])
    else:
        pool.insert(2, [f"Numero carta: **** **** **** {inst.cc_number[-4:]}"])

    pool += [
        [f"Movimenti al 28/02/2026"],
        [f"Filiale: SEDE CENTRALE"],
        [],
        [f"Saldo iniziale: {_format_amount_italian(random.uniform(1000, 25000))} EUR"],
        [],
        [f"Codice cliente: {random.randint(100000, 999999)}"],
        [],
        [f"Divisa: EUR"],
        [],
    ]

    rows = []
    used = set()
    for _ in range(n_rows):
        if len(used) < len(pool):
            idx = random.randint(0, len(pool) - 1)
            while idx in used:
                idx = random.randint(0, len(pool) - 1)
            used.add(idx)
            rows.append(pool[idx])
        else:
            rows.append([])
    return rows


def _generate_footer(inst: Instrument, n_rows: int, n_data: int) -> list[list[str]]:
    """Generate footer rows."""
    if n_rows == 0:
        return []

    pool = [
        [f"Saldo finale: {_format_amount_italian(random.uniform(500, 30000))} EUR"],
        [f"Totale movimenti: {n_data}"],
        [],
        [f"Documento generato automaticamente - non necessita di firma"],
        [f"Banca {inst.banca} - Servizio Home Banking"],
    ]

    return pool[:n_rows]


# ── File writing ──────────────────────────────────────────────────────────

def _txn_to_row(
    txn: Transaction, schema: FileSchema, is_csv: bool
) -> list:
    """Convert a Transaction to a list of cell values following schema."""
    row = []

    # Date
    if is_csv:
        row.append(_format_date_cell(txn.date))
    else:
        row.append(datetime(txn.date.year, txn.date.month, txn.date.day))

    # Valuta date
    if schema.valuta_col:
        vd = txn.valuta_date or txn.date
        if is_csv:
            row.append(_format_date_cell(vd))
        else:
            row.append(datetime(vd.year, vd.month, vd.day))

    # Description
    row.append(txn.description)

    # Amount
    amount = txn.amount
    if schema.amount_format == "signed_single":
        if is_csv:
            row.append(_format_amount_italian(amount))
        else:
            row.append(round(amount, 2))
    elif schema.amount_format == "positive_only":
        # Credit card: positive = expense
        if is_csv:
            row.append(_format_amount_italian(abs(amount)))
        else:
            row.append(round(abs(amount), 2))
    elif schema.amount_format == "debit_credit_split":
        # Two columns, both positive
        if amount < 0:
            debit_val = abs(amount)
            credit_val = ""
        else:
            debit_val = ""
            credit_val = amount
        if is_csv:
            row.append(_format_amount_italian(debit_val) if debit_val != "" else "")
            row.append(_format_amount_italian(credit_val) if credit_val != "" else "")
        else:
            row.append(round(debit_val, 2) if debit_val != "" else None)
            row.append(round(credit_val, 2) if credit_val != "" else None)
    elif schema.amount_format == "debit_credit_signed":
        # Two columns, debit has negative sign
        if amount < 0:
            debit_val = amount  # negative
            credit_val = ""
        else:
            debit_val = ""
            credit_val = amount
        if is_csv:
            row.append(_format_amount_italian(debit_val) if debit_val != "" else "")
            row.append(_format_amount_italian(credit_val) if credit_val != "" else "")
        else:
            row.append(round(debit_val, 2) if debit_val != "" else None)
            row.append(round(credit_val, 2) if credit_val != "" else None)

    # Currency
    if schema.currency_col:
        row.append("EUR")

    return row


def write_csv(
    filepath: Path,
    schema: FileSchema,
    preheader: list[list[str]],
    transactions: list[Transaction],
    footer: list[list[str]],
    separator: str = ";",
) -> None:
    """Write a CSV file with preheader, header, data, and footer."""
    n_cols = len(schema.column_names)

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=separator, quoting=csv.QUOTE_MINIMAL)

        # Preheader
        for row in preheader:
            padded = row + [""] * (n_cols - len(row))
            writer.writerow(padded[:n_cols])

        # Column header
        writer.writerow(schema.column_names)

        # Data rows
        for txn in transactions:
            writer.writerow(_txn_to_row(txn, schema, is_csv=True))

        # Footer
        for row in footer:
            padded = row + [""] * (n_cols - len(row))
            writer.writerow(padded[:n_cols])


def write_xlsx(
    filepath: Path,
    schema: FileSchema,
    preheader: list[list[str]],
    transactions: list[Transaction],
    footer: list[list[str]],
    has_borders: bool = False,
) -> None:
    """Write an XLSX file with preheader, header, data, and footer.

    If has_borders=True, the table region (header + data rows) is enclosed
    in a bordered rectangle mimicking real Italian bank XLSX exports.
    Preheader and footer remain outside the border, testing that
    detect_bordered_region() correctly identifies the table bounds.
    """
    if not HAS_OPENPYXL:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimenti"

    # Border styles for bordered files (thin = most common in Italian bank exports)
    if has_borders:
        _thin = Side(style="thin")
        _border_all = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
        _header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    else:
        _border_all = None
        _header_fill = None

    row_num = 1
    n_cols = len(schema.column_names)

    # Preheader (always outside border)
    for row in preheader:
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_num, column=col_idx, value=val)
        row_num += 1

    # Track table region for border application
    table_start_row = row_num

    # Column header
    for col_idx, name in enumerate(schema.column_names, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=name)
        cell.font = Font(bold=True)
        if has_borders:
            cell.border = _border_all
            cell.fill = _header_fill
    row_num += 1

    # Data rows
    for txn in transactions:
        values = _txn_to_row(txn, schema, is_csv=False)
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            if has_borders:
                cell.border = _border_all
        row_num += 1

    table_end_row = row_num - 1  # last data row

    # Footer (always outside border)
    for row in footer:
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_num, column=col_idx, value=val)
        row_num += 1

    # Auto-width columns
    for col_idx in range(1, n_cols + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 25

    wb.save(filepath)


# ── File plan ─────────────────────────────────────────────────────────────

@dataclass
class FilePlan:
    filename: str
    account_id: str
    format: Literal["csv", "xlsx"]
    separator: str  # "," or ";" for csv, "" for xlsx
    n_data_rows: int
    n_header_rows: int
    n_footer_rows: int
    schema_index: int  # for unique column combinations
    size_label: str  # S, M, L, V (variant)
    has_borders: bool = False  # True → add cell borders to XLSX table region


def _build_file_plans() -> list[FilePlan]:
    """Build the complete list of files to generate."""
    plans: list[FilePlan] = []
    schema_idx = 0

    sizes = {
        "S": 30,
        "M": None,  # random 50-150
        "L": 300,
    }

    for inst in INSTRUMENTS:
        for size_label, n_rows_or_none in sizes.items():
            n_data = n_rows_or_none if n_rows_or_none else random.randint(50, 150)

            # Alternate between csv and xlsx
            if schema_idx % 3 == 0:
                fmt = "csv"
                sep = ";"
            elif schema_idx % 3 == 1:
                fmt = "xlsx"
                sep = ""
            else:
                fmt = "csv"
                sep = ","

            # Vary preheader and footer
            n_header = random.choice([0, 0, 3, 5, 7, 10, 12])
            n_footer = random.choice([0, 0, 0, 2, 3, 4])

            # ~50% of XLSX files get borders (mimics real Italian bank exports)
            borders = fmt == "xlsx" and schema_idx % 2 == 1

            fname = f"{inst.id}_{size_label}_{schema_idx:03d}.{fmt}"
            plans.append(FilePlan(
                filename=fname,
                account_id=inst.id,
                format=fmt,
                separator=sep,
                n_data_rows=n_data,
                n_header_rows=n_header,
                n_footer_rows=n_footer,
                schema_index=schema_idx,
                size_label=size_label,
                has_borders=borders,
            ))
            schema_idx += 1

    # Add ~23 variant files for more coverage
    # "borders" key: True = bordered table (only XLSX), tests border detection
    variant_configs: list[dict] = [
        # Different formats for CC-1
        {"account_id": "CC-1", "fmt": "xlsx", "sep": "", "n_data": 80,
         "n_header": 15, "n_footer": 5, "label": "V", "borders": True},
        {"account_id": "CC-1", "fmt": "csv", "sep": ",", "n_data": 45,
         "n_header": 0, "n_footer": 0, "label": "V", "borders": False},
        {"account_id": "CC-1", "fmt": "csv", "sep": ";", "n_data": 100,
         "n_header": 8, "n_footer": 3, "label": "V", "borders": False},
        # CC-2 variants
        {"account_id": "CC-2", "fmt": "xlsx", "sep": "", "n_data": 60,
         "n_header": 4, "n_footer": 2, "label": "V", "borders": True},
        {"account_id": "CC-2", "fmt": "csv", "sep": ",", "n_data": 30,
         "n_header": 0, "n_footer": 0, "label": "V", "borders": False},
        # CC-3 variants
        {"account_id": "CC-3", "fmt": "csv", "sep": ";", "n_data": 50,
         "n_header": 6, "n_footer": 1, "label": "V", "borders": False},
        {"account_id": "CC-3", "fmt": "xlsx", "sep": "", "n_data": 90,
         "n_header": 10, "n_footer": 4, "label": "V", "borders": True},
        # CRED-1 variants
        {"account_id": "CRED-1", "fmt": "csv", "sep": ";", "n_data": 70,
         "n_header": 3, "n_footer": 2, "label": "V", "borders": False},
        {"account_id": "CRED-1", "fmt": "xlsx", "sep": "", "n_data": 40,
         "n_header": 0, "n_footer": 0, "label": "V", "borders": False},
        # CRED-2 variants
        {"account_id": "CRED-2", "fmt": "xlsx", "sep": "", "n_data": 55,
         "n_header": 5, "n_footer": 3, "label": "V", "borders": True},
        {"account_id": "CRED-2", "fmt": "csv", "sep": ",", "n_data": 120,
         "n_header": 12, "n_footer": 4, "label": "V", "borders": False},
        # RIC-1 (Sofia PostePay) variants
        {"account_id": "RIC-1", "fmt": "csv", "sep": ";", "n_data": 35,
         "n_header": 2, "n_footer": 1, "label": "V", "borders": False},
        {"account_id": "RIC-1", "fmt": "xlsx", "sep": "", "n_data": 85,
         "n_header": 7, "n_footer": 2, "label": "V", "borders": True},
        # RIC-2 (Marco Hype) variants
        {"account_id": "RIC-2", "fmt": "csv", "sep": ",", "n_data": 45,
         "n_header": 0, "n_footer": 0, "label": "V", "borders": False},
        {"account_id": "RIC-2", "fmt": "xlsx", "sep": "", "n_data": 65,
         "n_header": 9, "n_footer": 3, "label": "V", "borders": True},
        # RIC-3 (Laura Revolut) variants
        {"account_id": "RIC-3", "fmt": "xlsx", "sep": "", "n_data": 50,
         "n_header": 0, "n_footer": 0, "label": "V", "borders": False},
        {"account_id": "RIC-3", "fmt": "csv", "sep": ";", "n_data": 110,
         "n_header": 11, "n_footer": 4, "label": "V", "borders": False},
        # RISP-1 variants
        {"account_id": "RISP-1", "fmt": "csv", "sep": ",", "n_data": 40,
         "n_header": 5, "n_footer": 2, "label": "V", "borders": False},
        {"account_id": "RISP-1", "fmt": "xlsx", "sep": "", "n_data": 75,
         "n_header": 13, "n_footer": 5, "label": "V", "borders": True},
        # Extra edge cases: minimal / maximal headers
        {"account_id": "CC-1", "fmt": "csv", "sep": ";", "n_data": 30,
         "n_header": 1, "n_footer": 0, "label": "V", "borders": False},
        {"account_id": "CRED-1", "fmt": "csv", "sep": ",", "n_data": 30,
         "n_header": 14, "n_footer": 5, "label": "V", "borders": False},
        {"account_id": "CC-2", "fmt": "xlsx", "sep": "", "n_data": 30,
         "n_header": 0, "n_footer": 5, "label": "V", "borders": True},
        {"account_id": "CC-3", "fmt": "csv", "sep": ";", "n_data": 150,
         "n_header": 0, "n_footer": 0, "label": "V", "borders": False},
    ]

    for vc in variant_configs:
        fname = f"{vc['account_id']}_{vc['label']}_{schema_idx:03d}.{vc['fmt']}"
        plans.append(FilePlan(
            filename=fname,
            account_id=vc["account_id"],
            format=vc["fmt"],
            separator=vc["sep"],
            n_data_rows=vc["n_data"],
            n_header_rows=vc["n_header"],
            n_footer_rows=vc["n_footer"],
            schema_index=schema_idx,
            size_label=vc["label"],
            has_borders=vc.get("borders", False),
        ))
        schema_idx += 1

    return plans


# ── Manifest ──────────────────────────────────────────────────────────────

@dataclass
class ManifestRow:
    filename: str
    doc_type: str
    account_id: str
    bank: str
    format: str
    separator: str
    n_rows_total: int
    n_header_rows: int
    n_data_rows: int
    n_footer_rows: int
    amount_format: str
    has_debit_credit_split: str
    column_names: str
    total_income: float = 0.0
    total_expense: float = 0.0
    n_internal_transfers: int = 0
    n_income_rows: int = 0
    n_expense_rows: int = 0
    has_borders: str = "false"  # "true" if XLSX was generated with bordered table


def write_manifest(rows: list[ManifestRow], path: Path) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "filename", "doc_type", "account_id", "bank", "format",
            "separator", "n_rows_total", "n_header_rows", "n_data_rows",
            "n_footer_rows", "amount_format", "has_debit_credit_split",
            "column_names", "total_income", "total_expense",
            "n_internal_transfers", "n_income_rows", "n_expense_rows",
            "has_borders",
        ])
        for r in rows:
            writer.writerow([
                r.filename, r.doc_type, r.account_id, r.bank, r.format,
                r.separator, r.n_rows_total, r.n_header_rows, r.n_data_rows,
                r.n_footer_rows, r.amount_format, r.has_debit_credit_split,
                r.column_names, f"{r.total_income:.2f}",
                f"{r.total_expense:.2f}", r.n_internal_transfers,
                r.n_income_rows, r.n_expense_rows, r.has_borders,
            ])


# ── Ground truth generation ───────────────────────────────────────────────

def _classify_tx(txn: Transaction) -> str:
    """Classify a transaction into tx_type based on amount and giroconto flag."""
    if txn.is_giroconto:
        return "internal_out" if txn.amount < 0 else "internal_in"
    return "expense" if txn.amount < 0 else "income"


def write_ground_truth(
    filepath: Path,
    transactions: list[Transaction],
) -> dict:
    """Write a .expected.csv ground truth file alongside the generated file.

    Returns a summary dict with totals for the manifest.
    """
    expected_path = filepath.parent / (filepath.stem + ".expected.csv")

    total_income = 0.0
    total_expense = 0.0
    n_internal = 0
    n_income = 0
    n_expense = 0

    with open(expected_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "row_num", "date", "amount", "description_raw",
            "tx_type", "is_internal_transfer", "expected_category",
        ])
        for row_num, txn in enumerate(transactions, start=1):
            tx_type = _classify_tx(txn)
            canonical_amount = round(txn.amount, 2)
            is_internal = "true" if txn.is_giroconto else "false"

            writer.writerow([
                row_num,
                txn.date.strftime("%Y-%m-%d"),
                f"{canonical_amount:.2f}",
                txn.description,
                tx_type,
                is_internal,
                txn.expected_category,
            ])

            if canonical_amount > 0:
                total_income += canonical_amount
                n_income += 1
            elif canonical_amount < 0:
                total_expense += abs(canonical_amount)
                n_expense += 1

            if txn.is_giroconto:
                n_internal += 1

    return {
        "total_income": round(total_income, 2),
        "total_expense": round(total_expense, 2),
        "n_internal_transfers": n_internal,
        "n_income_rows": n_income,
        "n_expense_rows": n_expense,
    }


# ── Main generation ───────────────────────────────────────────────────────

def main() -> None:
    # Re-seed for determinism
    random.seed(42)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    plans = _build_file_plans()
    manifest_rows: list[ManifestRow] = []

    total_data_rows = 0
    n_csv = 0
    n_xlsx = 0

    for plan in plans:
        inst = INSTRUMENT_MAP[plan.account_id]
        is_cc = inst.doc_type == "credit_card"
        schema = _make_schema(plan.schema_index, is_credit_card=is_cc)

        # Generate transactions
        txns = generate_transactions(plan.account_id, plan.n_data_rows)

        # Generate preheader and footer
        preheader = _generate_preheader(inst, plan.n_header_rows)
        footer = _generate_footer(inst, plan.n_footer_rows, plan.n_data_rows)

        filepath = OUTPUT_DIR / plan.filename

        if plan.format == "csv":
            write_csv(filepath, schema, preheader, txns, footer,
                      separator=plan.separator)
            n_csv += 1
        else:
            write_xlsx(filepath, schema, preheader, txns, footer,
                       has_borders=plan.has_borders)
            n_xlsx += 1

        # Write per-file ground truth and collect summary stats
        gt_summary = write_ground_truth(filepath, txns)

        total_data_rows += plan.n_data_rows

        # +1 for the column header row
        n_total = plan.n_header_rows + 1 + plan.n_data_rows + plan.n_footer_rows

        manifest_rows.append(ManifestRow(
            filename=plan.filename,
            doc_type=inst.doc_type,
            account_id=plan.account_id,
            bank=inst.banca,
            format=plan.format,
            separator=plan.separator,
            n_rows_total=n_total,
            n_header_rows=plan.n_header_rows,
            n_data_rows=plan.n_data_rows,
            n_footer_rows=plan.n_footer_rows,
            amount_format=schema.amount_format,
            has_debit_credit_split="true" if schema.has_debit_credit_split else "false",
            column_names="|".join(schema.column_names),
            total_income=gt_summary["total_income"],
            total_expense=gt_summary["total_expense"],
            n_internal_transfers=gt_summary["n_internal_transfers"],
            n_income_rows=gt_summary["n_income_rows"],
            n_expense_rows=gt_summary["n_expense_rows"],
            has_borders="true" if plan.has_borders else "false",
        ))

    # Write manifest
    manifest_path = OUTPUT_DIR / "manifest.csv"
    write_manifest(manifest_rows, manifest_path)

    # Count ground truth files
    n_gt = sum(1 for _ in OUTPUT_DIR.glob("*.expected.csv"))

    # Print summary
    print("=" * 60)
    print("  Spendify — Synthetic Movements File Generator")
    print("=" * 60)
    print(f"  Output directory : {OUTPUT_DIR}")
    print(f"  Files generated  : {len(plans)}")
    print(f"    CSV files      : {n_csv}")
    print(f"    XLSX files     : {n_xlsx}")
    print(f"    Ground truth   : {n_gt}")
    print(f"  Total data rows  : {total_data_rows:,}")
    print(f"  Manifest         : {manifest_path}")
    print()

    # Per-account summary
    by_account: dict[str, list[FilePlan]] = {}
    for p in plans:
        by_account.setdefault(p.account_id, []).append(p)

    print(f"  {'Account':<10} {'Files':>5}  {'Data rows':>10}  Sizes")
    print(f"  {'-'*10} {'-'*5}  {'-'*10}  {'-'*20}")
    for acc_id in [i.id for i in INSTRUMENTS]:
        acc_plans = by_account.get(acc_id, [])
        n_files = len(acc_plans)
        n_rows = sum(p.n_data_rows for p in acc_plans)
        sizes_str = ", ".join(
            f"{p.size_label}={p.n_data_rows}" for p in acc_plans
        )
        print(f"  {acc_id:<10} {n_files:>5}  {n_rows:>10,}  {sizes_str}")

    print()
    print(f"  Amount formats used:")
    fmt_counts: dict[str, int] = {}
    for mr in manifest_rows:
        fmt_counts[mr.amount_format] = fmt_counts.get(mr.amount_format, 0) + 1
    for fmt_name, cnt in sorted(fmt_counts.items()):
        print(f"    {fmt_name:<25} {cnt:>3} files")

    print()
    n_bordered = sum(1 for p in plans if p.has_borders)
    print(f"  Header rows range : {min(p.n_header_rows for p in plans)}-{max(p.n_header_rows for p in plans)}")
    print(f"  Footer rows range : {min(p.n_footer_rows for p in plans)}-{max(p.n_footer_rows for p in plans)}")
    print(f"  XLSX with borders : {n_bordered} / {n_xlsx}")
    print("=" * 60)


if __name__ == "__main__":
    main()
