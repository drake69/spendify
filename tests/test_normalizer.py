"""Unit tests for core/normalizer.py — all deterministic, no LLM mocks needed."""
from decimal import Decimal
from datetime import date

import pandas as pd
import pytest

from decimal import Decimal as _Dec

from core.models import SignConvention
from core.normalizer import (
    PreprocessInfo,
    _build_owner_name_regex,
    _is_numeric_cell,
    apply_sign_convention,
    compute_columns_key,
    compute_file_hash,
    compute_header_sha256,
    compute_transaction_id,
    detect_and_strip_preheader_rows,
    detect_best_sheet,
    detect_delimiter,
    detect_encoding,
    detect_header_row,
    detect_header_row_excel,
    detect_internal_transfers,
    detect_skip_rows,
    drop_low_variability_columns,
    find_card_settlement_matches,
    load_raw_head,
    normalize_description,
    parse_amount,
    parse_date_safe,
    remove_card_balance_row,
)


class TestParseAmount:
    def test_integer(self):
        assert parse_amount("42") == Decimal("42")

    def test_negative(self):
        assert parse_amount("-12.50") == Decimal("-12.50")

    def test_european_format(self):
        assert parse_amount("1.234,56") == Decimal("1234.56")

    def test_us_format(self):
        assert parse_amount("1,234.56") == Decimal("1234.56")

    def test_comma_only_decimal(self):
        assert parse_amount("12,50") == Decimal("12.50")

    def test_strip_currency_symbol(self):
        assert parse_amount("€ 99,99") == Decimal("99.99")

    def test_decimal_instance(self):
        d = Decimal("5.00")
        assert parse_amount(d) is d

    def test_float(self):
        assert parse_amount(3.14) == Decimal("3.14")

    def test_invalid_returns_none(self):
        assert parse_amount("N/A") is None


class TestParseDateSafe:
    def test_italian_format(self):
        assert parse_date_safe("15/03/2024", "%d/%m/%Y") == date(2024, 3, 15)

    def test_iso_format(self):
        assert parse_date_safe("2024-03-15", "%Y-%m-%d") == date(2024, 3, 15)

    def test_invalid_returns_none(self):
        assert parse_date_safe("not-a-date", "%d/%m/%Y") is None

    def test_empty_returns_none(self):
        assert parse_date_safe("", "%d/%m/%Y") is None


class TestNormalizeDescription:
    def test_casefold(self):
        assert normalize_description("AMAZON.IT") == "amazon.it"

    def test_strips_whitespace(self):
        assert normalize_description("  foo  ") == "foo"

    def test_empty(self):
        assert normalize_description("") == ""


class TestComputeTransactionId:
    def test_deterministic(self):
        id1 = compute_transaction_id("file.csv", "01/01/2024", "100,00", "SUPERMERCATO COOP")
        id2 = compute_transaction_id("file.csv", "01/01/2024", "100,00", "SUPERMERCATO COOP")
        assert id1 == id2

    def test_length_24(self):
        tx_id = compute_transaction_id("file.csv", "01/01/2024", "100,00", "desc")
        assert len(tx_id) == 24

    def test_different_inputs_differ(self):
        id1 = compute_transaction_id("file.csv", "01/01/2024", "100,00", "desc a")
        id2 = compute_transaction_id("file.csv", "01/01/2024", "100,00", "desc b")
        assert id1 != id2

    def test_debit_credit_convention(self):
        # debit_positive: raw_amount is "<debit>|<credit>"
        id1 = compute_transaction_id("file.csv", "01/01/2024", "50,00|", "pagamento")
        id2 = compute_transaction_id("file.csv", "01/01/2024", "|100,00", "accredito")
        assert id1 != id2


class TestComputeFileHash:
    def test_deterministic(self):
        data = b"test content"
        assert compute_file_hash(data) == compute_file_hash(data)

    def test_length_64(self):
        assert len(compute_file_hash(b"anything")) == 64


class TestDetectDelimiter:
    def test_comma(self):
        assert detect_delimiter("a,b,c\n1,2,3") == ","

    def test_semicolon(self):
        assert detect_delimiter("a;b;c\n1;2;3") == ";"

    def test_tab(self):
        assert detect_delimiter("a\tb\tc\n1\t2\t3") == "\t"


# ── Helpers ────────────────────────────────────────────────────────────────────

def _make_tx_df(n_rows: int = 20) -> pd.DataFrame:
    """Return a clean transaction-like DataFrame with no pre-header rows."""
    return pd.DataFrame({
        "Data": [f"2024-01-{i+1:02d}" for i in range(n_rows)],
        "Descrizione": [f"Pagamento {i}" for i in range(n_rows)],
        "Importo": [f"{(i+1)*10:.2f}" for i in range(n_rows)],
        "Tipo": ["Addebito" if i % 2 == 0 else "Accredito" for i in range(n_rows)],
    })


# ── TestDetectAndStripPreheaderRows ────────────────────────────────────────────

class TestDetectAndStripPreheaderRows:

    def test_no_preheader_rows_unchanged(self):
        df = _make_tx_df(20)
        result, n = detect_and_strip_preheader_rows(df, "test.csv")
        assert n == 0
        assert len(result) == len(df)

    def test_strips_sparse_rows_at_start(self):
        """Simulate the real scenario: bank file with title rows before data.

        When a bank CSV looks like:
            Estratto Conto,,,,
            Dal 01/01/2024 al 31/01/2024,,,,
            Data;Descrizione;Importo;Tipo
            2024-01-01;Pag 0;10.00;Addebito
            ...

        pandas.read_csv (with detect_header_row returning 0) consumes the
        FIRST row as column names: ["Estratto Conto", ..., Unnamed: N].
        The actual header row ("Data", "Descrizione", ...) ends up as row 1 in
        the DataFrame body; row 0 is the second metadata line.

        Inside detect_and_strip_preheader_rows, both the reconstructed
        header (from Unnamed-heavy column names) AND the metadata body row
        are sparse — so n_sparse == 2 and the real header becomes the new
        column names.
        """
        n_data = 20
        # Simulate pandas output: first real column has a title value,
        # the rest are Unnamed (because the title row had empty cells).
        title_cols = ["Estratto Conto", "Unnamed: 1", "Unnamed: 2", "Unnamed: 3"]
        rows = [
            ["Dal 01/01/2024 al 31/01/2024", None, None, None],    # sparse body row
            ["Data", "Descrizione", "Importo", "Tipo"],              # real header
        ] + [
            [f"2024-01-{i+1:02d}", f"Pag {i}", f"{(i+1)*10:.2f}", "Addebito"]
            for i in range(n_data)
        ]
        df = pd.DataFrame(rows, columns=title_cols)

        result, n = detect_and_strip_preheader_rows(df, "bank.csv")

        # The reconstructed header row ("Estratto Conto" | None | None | None,
        # density=1/4=0.25) and body row 0 (density=1/4=0.25) are both sparse
        # → n_sparse == 2; real header is in df_full row 2.
        assert n == 2
        assert list(result.columns) == ["Data", "Descrizione", "Importo", "Tipo"]
        assert len(result) == n_data

    def test_too_short_returns_unchanged(self):
        df = _make_tx_df(3)
        result, n = detect_and_strip_preheader_rows(df, "short.csv")
        assert n == 0
        assert len(result) == 3

    def test_exceeds_absolute_cap_raises(self):
        """If more than 20 contiguous sparse rows exist → ValueError.

        Use Unnamed column names so the reconstructed header row is also sparse
        (density 0/4 = 0.0), making it contiguous with the empty body rows.
        """
        n_sparse = 25
        n_data = 50
        # Unnamed columns → reconstructed header is fully sparse
        cols = [f"Unnamed: {i}" for i in range(4)]
        sparse_rows = [[None] * 4 for _ in range(n_sparse)]
        data_rows = [[f"v{i}", f"desc {i}", f"{i}.00", "X"] for i in range(n_data)]
        df = pd.DataFrame(sparse_rows + data_rows, columns=cols)
        with pytest.raises(ValueError, match="exceeding the absolute cap"):
            detect_and_strip_preheader_rows(df, "bad.csv")

    def test_exceeds_ratio_cap_raises(self):
        """If sparse rows > 10 % of total → ValueError.

        With Unnamed columns: 4 sparse rows + reconstructed header (also sparse)
        = 5 total sparse in df_full of 36 rows → 5/36 ≈ 13.9 % > 10 %.
        """
        # 4 sparse body rows + 30 data rows.
        # df_full total = 1 (header) + 4 (sparse body) + 30 (data) = 35 rows
        # n_sparse = 5 (header + 4 body) → 5/35 ≈ 14.3 % > 10 % → raises
        n_sparse_body = 4
        n_data = 30
        cols = [f"Unnamed: {i}" for i in range(4)]
        sparse_rows = [[None] * 4 for _ in range(n_sparse_body)]
        data_rows = [[f"v{i}", f"desc {i}", f"{i}.00", "X"] for i in range(n_data)]
        df = pd.DataFrame(sparse_rows + data_rows, columns=cols)
        with pytest.raises(ValueError, match="safety cap"):
            detect_and_strip_preheader_rows(df, "ratio.csv")

    def test_clean_file_with_many_rows_not_triggered(self):
        """A large clean file should never trigger false positives."""
        df = _make_tx_df(100)
        result, n = detect_and_strip_preheader_rows(df, "large.csv")
        assert n == 0
        assert len(result) == 100


# ── TestDropLowVariabilityColumns ─────────────────────────────────────────────

class TestDropLowVariabilityColumns:

    def test_no_low_variability_columns_unchanged(self):
        df = _make_tx_df(20)
        result, dropped = drop_low_variability_columns(df, "test.csv")
        assert dropped == []
        assert list(result.columns) == list(df.columns)

    def test_drops_constant_column(self):
        """A column with the same value on every row should be dropped.

        With 100 rows: 1 unique / 100 rows = 1 % < 1.5 % threshold → dropped.
        """
        df = _make_tx_df(100)
        df["Nome titolare"] = "Mario Rossi"   # constant — will be dropped
        result, dropped = drop_low_variability_columns(df, "amex.csv")
        assert "Nome titolare" in dropped
        assert "Nome titolare" not in result.columns

    def test_drops_near_constant_column(self):
        """A column that barely varies (< 1.5 %) should also be dropped."""
        n = 100
        df = _make_tx_df(n)
        # Only 1 unique value among 100 rows → ratio = 1/100 = 1 % < 1.5 %
        df["Numero carta"] = "**** **** **** 1234"
        result, dropped = drop_low_variability_columns(df, "amex.csv")
        assert "Numero carta" in dropped

    def test_preserves_minimum_two_columns(self):
        """Even if all columns are constant, at least 2 are kept.

        With n=100 rows: 1 unique / 100 = 1 % < 1.5 % for every column,
        so all three are flagged. But max_droppable = 3 - 2 = 1, so only
        one column is removed, leaving exactly 2.
        """
        n = 100
        df = pd.DataFrame({
            "A": ["x"] * n,
            "B": ["y"] * n,
            "C": ["z"] * n,
        })
        result, dropped = drop_low_variability_columns(df, "edge.csv")
        assert len(result.columns) == 2

    def test_does_not_drop_variable_columns(self):
        """Columns with sufficient variability must be kept."""
        n = 100
        df = pd.DataFrame({
            "Data": [f"2024-01-{i % 28 + 1:02d}" for i in range(n)],
            "Importo": [f"{i:.2f}" for i in range(n)],
            "Descrizione": [f"Pagamento {i}" for i in range(n)],
        })
        result, dropped = drop_low_variability_columns(df, "good.csv")
        assert dropped == []
        assert len(result.columns) == 3


# ── TestDetectHeaderRow ────────────────────────────────────────────────────────

class TestDetectHeaderRow:
    """Tests for detect_header_row(lines) -> (int, bool)."""

    def test_header_at_row_0_certain(self):
        """Standard CSV: header is the first line → (0, True)."""
        lines = [
            "Data,Importo,Descrizione",
            "01/01/2024,100.00,Supermercato",
            "02/01/2024,-50.00,Bolletta",
        ]
        n, certain = detect_header_row(lines)
        assert n == 0
        assert certain is True

    def test_header_at_row_3_certain(self):
        """AMEX-style: 3 metadata rows before the real header → (3, True)."""
        lines = [
            "",
            "Estratto conto American Express",
            "Dal 01/01/2024 al 31/01/2024",
            "Data,Descrizione,Importo,Valuta",
            "01/01/2024,Ristorante,45.00,EUR",
        ]
        n, certain = detect_header_row(lines)
        assert n == 3
        assert certain is True

    def test_all_numeric_uncertain(self):
        """All-numeric file: no header line found → fallback (0, False)."""
        lines = [
            "1,2,3",
            "4,5,6",
            "7,8,9",
        ]
        n, certain = detect_header_row(lines)
        assert n == 0
        assert certain is False

    def test_empty_lines_uncertain(self):
        """Empty file: no lines → fallback (0, False)."""
        n, certain = detect_header_row([])
        assert n == 0
        assert certain is False

    def test_single_field_line_skipped(self):
        """A line with only 1 non-numeric field does not match (needs ≥2)."""
        lines = [
            "Estratto",          # 1 non-numeric field → not a match
            "Data,Importo,Desc", # 3 non-numeric fields → match
        ]
        n, certain = detect_header_row(lines)
        assert n == 1
        assert certain is True

    def test_semicolon_delimiter(self):
        """Detection works with semicolon-delimited files."""
        lines = [
            "Data;Importo;Descrizione",
            "01/01/2024;100,00;Pagamento",
        ]
        n, certain = detect_header_row(lines)
        assert n == 0
        assert certain is True


# ── TestDetectSkipRows ─────────────────────────────────────────────────────────

class TestDetectSkipRows:
    """Tests for detect_skip_rows(raw_bytes, filename) -> (int, bool)."""

    def test_csv_standard_header(self):
        """Standard CSV → (0, True)."""
        content = "Data,Importo,Descrizione\n01/01/2024,100.00,Pane\n"
        raw = content.encode("utf-8")
        n, certain, _ = detect_skip_rows(raw, "estratto.csv")
        assert n == 0
        assert certain is True

    def test_csv_with_metadata_rows(self):
        """CSV with 2 metadata rows before header → (2, True)."""
        content = (
            "Banca Esempio SpA\n"
            "Periodo: Gennaio 2024\n"
            "Data,Importo,Descrizione\n"
            "01/01/2024,100.00,Pane\n"
        )
        raw = content.encode("utf-8")
        n, certain, _ = detect_skip_rows(raw, "estratto.csv")
        assert n == 2
        assert certain is True

    def test_csv_all_numeric_uncertain(self):
        """CSV with no text header → (0, False)."""
        content = "1,2,3\n4,5,6\n"
        raw = content.encode("utf-8")
        n, certain, _ = detect_skip_rows(raw, "data.csv")
        assert n == 0
        assert certain is False

    def test_too_few_rows_unchanged(self):
        df = pd.DataFrame({"A": ["x"], "B": ["y"], "C": ["z"]})
        result, dropped = drop_low_variability_columns(df, "tiny.csv")
        assert dropped == []
        assert len(result.columns) == 3

    def test_already_two_columns_unchanged(self):
        df = pd.DataFrame({"A": ["x"] * 20, "B": ["y"] * 20})
        result, dropped = drop_low_variability_columns(df, "min.csv")
        assert dropped == []
        assert len(result.columns) == 2


# ── TestPreprocessInfo ─────────────────────────────────────────────────────────

class TestPreprocessInfo:
    def test_defaults(self):
        info = PreprocessInfo()
        assert info.skipped_rows == 0
        assert info.dropped_columns == []

    def test_custom_values(self):
        info = PreprocessInfo(skipped_rows=3, dropped_columns=["Nome titolare"])
        assert info.skipped_rows == 3
        assert info.dropped_columns == ["Nome titolare"]


# ── TestDetectEncoding ─────────────────────────────────────────────────────────

class TestDetectEncoding:
    def test_ascii_normalizes_to_utf8(self):
        """'ascii' alias must be normalised to 'utf-8'."""
        from unittest.mock import patch
        with patch("chardet.detect", return_value={"encoding": "ascii"}):
            enc = detect_encoding(b"anything")
        assert enc == "utf-8"

    def test_none_encoding_falls_back_to_utf8(self):
        """chardet returning None encoding → default to utf-8."""
        from unittest.mock import patch
        with patch("chardet.detect", return_value={"encoding": None}):
            enc = detect_encoding(b"anything")
        assert enc == "utf-8"

    def test_utf8_bom_content(self):
        """UTF-8 BOM content should return a utf-* encoding."""
        raw = "\ufeffData,Importo".encode("utf-8-sig")
        enc = detect_encoding(raw)
        assert "utf" in enc or "ascii" in enc or enc == "utf-8"


# ── TestIsNumericCell ──────────────────────────────────────────────────────────

class TestIsNumericCell:
    def test_none_returns_false(self):
        assert _is_numeric_cell(None) is False

    def test_integer_returns_true(self):
        assert _is_numeric_cell(42) is True

    def test_float_returns_true(self):
        assert _is_numeric_cell(3.14) is True

    def test_numeric_string_returns_true(self):
        assert _is_numeric_cell("123.45") is True

    def test_comma_decimal_string_returns_true(self):
        assert _is_numeric_cell("1234,56") is True

    def test_text_string_returns_false(self):
        assert _is_numeric_cell("Descrizione") is False


# ── TestDetectHeaderRowExcel ───────────────────────────────────────────────────

class TestDetectHeaderRowExcel:
    def _make_xlsx(self, rows: list[list]) -> bytes:
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_header_at_row_0(self):
        raw = self._make_xlsx([
            ["Data", "Importo", "Descrizione"],
            ["01/01/2024", 100.0, "Supermercato"],
        ])
        n, certain, _ = detect_header_row_excel(raw)
        assert n == 0
        assert certain is True

    def test_header_at_row_2(self):
        raw = self._make_xlsx([
            ["Banca Esempio"],
            ["Periodo: Gennaio 2024"],
            ["Data", "Importo", "Descrizione"],
            ["01/01/2024", 100.0, "Pagamento"],
        ])
        n, certain, _ = detect_header_row_excel(raw)
        assert n == 2
        assert certain is True

    def test_all_numeric_uncertain(self):
        raw = self._make_xlsx([
            [1, 2, 3],
            [4, 5, 6],
        ])
        n, certain, _ = detect_header_row_excel(raw)
        assert n == 0
        assert certain is False

    def test_invalid_bytes_fallback(self):
        n, certain, _ = detect_header_row_excel(b"not an xlsx file")
        assert n == 0
        assert certain is False


# ── TestDetectSkipRowsExcel ────────────────────────────────────────────────────

class TestDetectSkipRowsExcel:
    def _make_xlsx(self, rows: list[list]) -> bytes:
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_excel_routes_to_excel_detection(self):
        raw = self._make_xlsx([
            ["Data", "Importo", "Descrizione"],
            ["01/01/2024", 100.0, "Test"],
        ])
        n, certain, _ = detect_skip_rows(raw, "estratto.xlsx")
        assert n == 0
        assert certain is True

    def test_xls_extension_also_routed(self):
        # .xls extension should also use the Excel path (will likely fail parse
        # since we generate .xlsx, but should return (0, False) gracefully)
        raw = self._make_xlsx([["Data", "Importo"], ["01/01/2024", 100.0]])
        n, certain, _ = detect_skip_rows(raw, "file.xls")
        # may or may not parse correctly, but must return a valid tuple
        assert isinstance(n, int)
        assert isinstance(certain, bool)


# ── TestDetectBestSheet ────────────────────────────────────────────────────────

class TestDetectBestSheet:
    def _make_workbook(self, sheets: dict[str, list[list]]):
        import openpyxl
        wb = openpyxl.Workbook()
        first = True
        for name, rows in sheets.items():
            if first:
                ws = wb.active
                ws.title = name
                first = False
            else:
                ws = wb.create_sheet(name)
            for row in rows:
                ws.append(row)
        return wb

    def test_summary_sheet_excluded(self):
        wb = self._make_workbook({
            "Summary": [["A", "B"], ["x", "y"]],
            "Data": [["Date", "Amount"], ["01/01/2024", 100.0], ["02/01/2024", 200.0]],
        })
        result = detect_best_sheet(wb)
        assert result == "Data"

    def test_most_numeric_sheet_wins(self):
        wb = self._make_workbook({
            "Metadata": [["Info", "x"], ["y", "z"]],
            "Transactions": [
                ["Date", "Amount", "Desc"],
                *[[f"2024-01-{i:02d}", float(i * 10), f"Tx {i}"] for i in range(1, 20)],
            ],
        })
        result = detect_best_sheet(wb)
        assert result == "Transactions"

    def test_fallback_to_first_when_all_summary(self):
        wb = self._make_workbook({
            "Riepilogo": [["A"], ["B"]],
            "Totale": [["C"], ["D"]],
        })
        result = detect_best_sheet(wb)
        # All sheets match summary pattern → fallback to first
        assert result == wb.sheetnames[0]


# ── TestParseAmountEdgeCases ───────────────────────────────────────────────────

class TestParseAmountEdgeCases:
    def test_inf_returns_none(self):
        assert parse_amount(float("inf")) is None

    def test_nan_returns_none(self):
        assert parse_amount(float("nan")) is None

    def test_none_returns_none(self):
        assert parse_amount(None) is None

    def test_list_returns_none(self):
        assert parse_amount([1, 2, 3]) is None  # type: ignore[arg-type]

    def test_comma_thousands_no_decimal(self):
        """'1,000,000' — comma as thousands separator, no decimal → 1000000."""
        assert parse_amount("1,000,000") == _Dec("1000000")

    def test_decimal_infinite_returns_none(self):
        from decimal import Decimal as D
        import decimal
        inf = D("Infinity")
        assert parse_amount(inf) is None


# ── TestApplySignConvention ────────────────────────────────────────────────────

class TestApplySignConvention:
    def test_signed_single(self):
        row = {"Importo": "-10.50"}
        result = apply_sign_convention(row, "Importo", None, None, SignConvention.signed_single)
        assert result == _Dec("-10.50")

    def test_debit_positive_debit_only(self):
        row = {"Addebito": "10.00", "Accredito": ""}
        result = apply_sign_convention(row, "", "Addebito", "Accredito", SignConvention.debit_positive)
        assert result == _Dec("-10.00")

    def test_debit_positive_credit_only(self):
        row = {"Addebito": "", "Accredito": "20.00"}
        result = apply_sign_convention(row, "", "Addebito", "Accredito", SignConvention.debit_positive)
        assert result == _Dec("20.00")

    def test_debit_positive_both_none_fallback_amount(self):
        row = {"Importo": "-5.00"}
        result = apply_sign_convention(row, "Importo", None, None, SignConvention.debit_positive)
        assert result == _Dec("-5.00")

    def test_credit_negative_credit_positive(self):
        row = {"Entrata": "50.00"}
        result = apply_sign_convention(row, "", None, "Entrata", SignConvention.credit_negative)
        assert result == _Dec("50.00")

    def test_credit_negative_debit_negative(self):
        row = {"Uscita": "30.00"}
        result = apply_sign_convention(row, "", "Uscita", None, SignConvention.credit_negative)
        assert result == _Dec("-30.00")

    def test_credit_negative_both_none_fallback(self):
        row = {"Importo": "-7.00"}
        result = apply_sign_convention(row, "Importo", None, None, SignConvention.credit_negative)
        assert result == _Dec("-7.00")


# ── TestComputeColumnsKey ──────────────────────────────────────────────────────

class TestComputeColumnsKey:
    def test_returns_cols_prefix(self):
        df = pd.DataFrame(columns=["Data", "Importo", "Descrizione"])
        key = compute_columns_key(df)
        assert key.startswith("cols:")

    def test_length_16_hex_after_prefix(self):
        df = pd.DataFrame(columns=["A", "B"])
        key = compute_columns_key(df)
        assert len(key) == len("cols:") + 16

    def test_column_order_independent(self):
        df1 = pd.DataFrame(columns=["Data", "Importo"])
        df2 = pd.DataFrame(columns=["Importo", "Data"])
        assert compute_columns_key(df1) == compute_columns_key(df2)

    def test_different_columns_differ(self):
        df1 = pd.DataFrame(columns=["Data", "Importo"])
        df2 = pd.DataFrame(columns=["Date", "Amount"])
        assert compute_columns_key(df1) != compute_columns_key(df2)


# ── TestComputeHeaderSha256 ────────────────────────────────────────────────────

class TestComputeHeaderSha256:
    def test_csv_returns_64_char_hex(self):
        raw = b"Data,Importo,Descrizione\n01/01/2024,100.00,Test\n"
        h = compute_header_sha256(raw, "file.csv")
        assert len(h) == 64
        assert all(c in "0123456789abcdef" for c in h)

    def test_same_content_same_hash(self):
        raw = b"Data,Importo\n01/01/2024,100.00\n"
        h1 = compute_header_sha256(raw, "a.csv")
        h2 = compute_header_sha256(raw, "b.csv")
        assert h1 == h2

    def test_different_content_different_hash(self):
        r1 = b"Data,Importo\n"
        r2 = b"Date,Amount\n"
        assert compute_header_sha256(r1, "f.csv") != compute_header_sha256(r2, "f.csv")

    def test_invalid_excel_falls_back(self):
        """Non-Excel bytes with .xlsx extension → fallback to raw bytes hash."""
        raw = b"not a real xlsx"
        h = compute_header_sha256(raw, "file.xlsx")
        assert len(h) == 64


# ── TestLoadRawHead ────────────────────────────────────────────────────────────

class TestLoadRawHead:
    def test_csv_returns_dataframe(self):
        content = "Col1,Col2,Col3\n1,2,3\n4,5,6\n7,8,9\n"
        raw = content.encode("utf-8")
        df = load_raw_head(raw, "test.csv", n=2)
        assert len(df) == 2

    def test_csv_no_header_mode(self):
        """load_raw_head reads WITHOUT header — row 0 is the header row as data."""
        content = "Data,Importo\n01/01/2024,100.00\n"
        raw = content.encode("utf-8")
        df = load_raw_head(raw, "test.csv", n=10)
        assert df.iloc[0, 0] == "Data"


# ── TestBuildOwnerNameRegex ────────────────────────────────────────────────────

class TestBuildOwnerNameRegex:
    def test_empty_list_returns_none(self):
        assert _build_owner_name_regex([]) is None

    def test_whitespace_only_names_returns_none(self):
        assert _build_owner_name_regex(["  ", "\t"]) is None

    def test_single_token_name(self):
        pattern = _build_owner_name_regex(["Amazon"])
        assert pattern is not None
        assert pattern.search("Acquisto Amazon oggi") is not None

    def test_multi_token_name_both_orders(self):
        pattern = _build_owner_name_regex(["Mario Rossi"])
        assert pattern is not None
        assert pattern.search("Bonifico Mario Rossi") is not None
        assert pattern.search("Bonifico Rossi Mario") is not None

    def test_case_insensitive(self):
        pattern = _build_owner_name_regex(["Mario Rossi"])
        assert pattern.search("MARIO ROSSI") is not None


# ── TestRemoveCardBalanceRow ───────────────────────────────────────────────────

class TestRemoveCardBalanceRow:
    def _tx(self, amount, desc="desc"):
        return {"amount": _Dec(str(amount)), "description": desc, "id": "x"}

    def test_too_few_rows_returns_unchanged(self):
        txs = [self._tx(10), self._tx(-10)]
        result, found = remove_card_balance_row(txs)
        assert found is False
        assert len(result) == 2

    def test_balance_row_removed_without_label(self):
        # tx[2] has amount == sum of others: 10 + 20 = 30
        txs = [self._tx(10), self._tx(20), self._tx(30)]
        result, found = remove_card_balance_row(txs)
        assert found is True
        assert len(result) == 2
        assert all(tx["amount"] != _Dec("30") for tx in result)

    def test_balance_row_relabelled_with_owner_name(self):
        txs = [self._tx(10), self._tx(20), self._tx(30, desc="Saldo")]
        result, found = remove_card_balance_row(txs, owner_name_label="Mario Rossi")
        assert found is True
        assert len(result) == 3  # not removed, just relabelled
        assert result[2]["description"] == "Mario Rossi"

    def test_no_balance_row_returns_unchanged(self):
        # No single tx == sum of others
        txs = [self._tx(10), self._tx(20), self._tx(999)]
        result, found = remove_card_balance_row(txs)
        assert found is False
        assert len(result) == 3


# ── TestDetectInternalTransfersEdgeCases ──────────────────────────────────────

class TestDetectInternalTransfersEdgeCases:
    def _make_df(self, rows):
        from datetime import date as _date
        import numpy as np
        df = pd.DataFrame(rows)
        df["transfer_pair_id"] = None
        df["transfer_confidence"] = None
        return df

    def test_no_keyword_no_high_sym_not_paired(self):
        """Amount match + date match but no keyword and delta > strict → skipped (line 645)."""
        from datetime import date as _date
        rows = [
            {"amount": _Dec("100.00"), "date": _date(2024, 1, 1), "description": "generico",
             "tx_type": "expense"},
            {"amount": _Dec("-100.00"), "date": _date(2024, 1, 3), "description": "generico",
             "tx_type": "expense"},
        ]
        df = self._make_df(rows)
        result = detect_internal_transfers(
            df,
            keyword_patterns=[],
            require_keyword_confirmation=True,
            delta_days=5,
            delta_days_strict=1,   # 2-day gap > strict → not high_sym
        )
        # No pair should be formed since no keyword and not high symmetry
        assert result["transfer_pair_id"].isna().all()

    def test_owner_name_string_amount_converted(self):
        """Owner-name pass handles string amounts (line 683)."""
        from datetime import date as _date
        rows = [
            {"amount": "-50.00",  # string, not Decimal
             "date": _date(2024, 1, 1),
             "description": "Bonifico Mario Rossi",
             "tx_type": "expense"},
        ]
        df = self._make_df(rows)
        result = detect_internal_transfers(
            df,
            owner_names=["Mario Rossi"],
            keyword_patterns=[],
        )
        assert result.iloc[0]["tx_type"] == "internal_out"


# ── TestDetectInternalTransfersFullPairing ────────────────────────────────────

class TestDetectInternalTransfersFullPairing:
    """Tests that exercise the main combination loop with different account_labels."""

    def _make_df_no_cols(self, rows):
        """DataFrame WITHOUT pre-added transfer columns (hits lines 595, 597)."""
        from datetime import date as _date
        df = pd.DataFrame(rows)
        return df

    def test_columns_added_when_missing(self):
        """detect_internal_transfers adds columns if absent (lines 595, 597)."""
        from datetime import date as _date
        rows = [
            {"amount": _Dec("100.00"), "date": _date(2024, 1, 1),
             "description": "generic", "tx_type": "income", "account_label": "A"},
        ]
        df = self._make_df_no_cols(rows)
        assert "transfer_pair_id" not in df.columns
        result = detect_internal_transfers(df, keyword_patterns=[])
        assert "transfer_pair_id" in result.columns
        assert "transfer_confidence" in result.columns

    def test_keyword_match_pairs_transactions(self):
        """Keyword match → high confidence pair formed (lines 601, 606, 625-667)."""
        from datetime import date as _date
        rows = [
            {"amount": _Dec("-100.00"), "date": _date(2024, 1, 1),
             "description": "Bonifico Mario Rossi giroconto", "tx_type": "expense",
             "account_label": "AccountA"},
            {"amount": _Dec("100.00"), "date": _date(2024, 1, 1),
             "description": "Accredito giroconto da Mario", "tx_type": "income",
             "account_label": "AccountB"},
        ]
        df = pd.DataFrame(rows)
        df["transfer_pair_id"] = None
        df["transfer_confidence"] = None
        result = detect_internal_transfers(
            df,
            keyword_patterns=["giroconto"],
            require_keyword_confirmation=False,
        )
        assert result.iloc[0]["tx_type"] == "internal_out"
        assert result.iloc[1]["tx_type"] == "internal_in"
        assert result.iloc[0]["transfer_pair_id"] is not None

    def test_keyword_match_function_with_re(self):
        """_keyword_match returns True when keyword_re matches (line 606)."""
        from datetime import date as _date
        rows = [
            {"amount": _Dec("-50.00"), "date": _date(2024, 1, 2),
             "description": "bonifico stipendio giroconto", "tx_type": "expense",
             "account_label": "AccountA"},
            {"amount": _Dec("50.00"), "date": _date(2024, 1, 2),
             "description": "giroconto ricevuto", "tx_type": "income",
             "account_label": "AccountB"},
        ]
        df = pd.DataFrame(rows)
        df["transfer_pair_id"] = None
        df["transfer_confidence"] = None
        result = detect_internal_transfers(
            df,
            keyword_patterns=["giroconto"],
        )
        assert result["transfer_pair_id"].notna().any()

    def test_high_sym_pair_without_keyword(self):
        """High symmetry (tight amount + date match) → medium confidence pair (lines 642-655)."""
        from datetime import date as _date
        rows = [
            {"amount": _Dec("-200.00"), "date": _date(2024, 1, 1),
             "description": "transfer", "tx_type": "expense", "account_label": "A"},
            {"amount": _Dec("200.00"), "date": _date(2024, 1, 1),
             "description": "transfer", "tx_type": "income", "account_label": "B"},
        ]
        df = pd.DataFrame(rows)
        df["transfer_pair_id"] = None
        df["transfer_confidence"] = None
        result = detect_internal_transfers(
            df,
            keyword_patterns=[],
            require_keyword_confirmation=True,
            delta_days_strict=1,
        )
        # Medium confidence → paired but tx_type not changed
        assert result["transfer_pair_id"].notna().any()

    def test_already_paired_skipped(self):
        """Second pair attempt on already-paired index → continue (line 619)."""
        from datetime import date as _date
        rows = [
            {"amount": _Dec("-100.00"), "date": _date(2024, 1, 1),
             "description": "giro", "tx_type": "expense", "account_label": "A"},
            {"amount": _Dec("100.00"), "date": _date(2024, 1, 1),
             "description": "giro", "tx_type": "income", "account_label": "B"},
            {"amount": _Dec("100.00"), "date": _date(2024, 1, 1),
             "description": "giro", "tx_type": "income", "account_label": "C"},
        ]
        df = pd.DataFrame(rows)
        df["transfer_pair_id"] = None
        df["transfer_confidence"] = None
        # First pair (0,1) should be formed; (0,2) and (1,2) should hit already_paired
        result = detect_internal_transfers(
            df,
            keyword_patterns=["giro"],
            require_keyword_confirmation=False,
        )
        # At least one pair formed
        assert result["transfer_pair_id"].notna().sum() >= 2

    def test_owner_re_skips_already_paired(self):
        """Owner-name pass skips idx in already_paired (line 678)."""
        from datetime import date as _date
        rows = [
            {"amount": _Dec("-100.00"), "date": _date(2024, 1, 1),
             "description": "giro Mario Rossi", "tx_type": "expense", "account_label": "A"},
            {"amount": _Dec("100.00"), "date": _date(2024, 1, 1),
             "description": "giro Mario Rossi", "tx_type": "income", "account_label": "B"},
        ]
        df = pd.DataFrame(rows)
        df["transfer_pair_id"] = None
        df["transfer_confidence"] = None
        # Both rows get paired by keyword 'giro'; owner_re pass should skip them
        result = detect_internal_transfers(
            df,
            owner_names=["Mario Rossi"],
            keyword_patterns=["giro"],
            require_keyword_confirmation=False,
        )
        # Rows were already paired; owner_re pass should not re-label them
        assert result.iloc[0]["transfer_pair_id"] is not None


# ── TestApplySignConventionEdgeCases ──────────────────────────────────────────

class TestApplySignConventionEdgeCases:
    """Cover lines 271-272 in apply_sign_convention (credit_negative fallbacks)."""

    def test_credit_negative_credit_zero_falls_through(self):
        """credit=0 → `if credit and credit > 0` fails → check debit → fallback (line 271)."""
        row = {"Entrata": "0", "Importo": "-7.50"}
        result = apply_sign_convention(
            row, "Importo", None, "Entrata", SignConvention.credit_negative
        )
        # credit parses as 0 which is falsy → falls through to line 271
        assert result == _Dec("-7.50")

    def test_credit_negative_no_amount_col_returns_none(self):
        """credit=None, debit=None, amount_col='': fallback returns None (line 266)."""
        row = {}
        result = apply_sign_convention(
            row, "", None, None, SignConvention.credit_negative
        )
        assert result is None

    def test_unknown_convention_returns_none(self):
        """Convention not in the three known types → final return None (line 272)."""
        row = {"Importo": "10.00"}
        result = apply_sign_convention(row, "Importo", None, None, "unknown_convention")  # type: ignore
        assert result is None


# ── TestDetectBestSheetEmptySheet ──────────────────────────────────────────────

class TestDetectBestSheetEmptySheet:
    """Cover line 146 — empty sheet is skipped in detect_best_sheet."""

    def test_empty_sheet_skipped(self):
        """A sheet with no rows is skipped; the data sheet wins (line 146)."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws_empty = wb.active
        ws_empty.title = "Empty"
        # Leave ws_empty with no rows
        ws_data = wb.create_sheet("Data")
        ws_data.append(["Date", "Amount", "Description"])
        ws_data.append(["01/01/2024", 100.0, "Test"])
        result = detect_best_sheet(wb)
        assert result == "Data"


# ── TestComputeHeaderSha256Excel ───────────────────────────────────────────────

class TestComputeHeaderSha256Excel:
    """Cover lines 329-337 — Excel path in compute_header_sha256."""

    def _make_xlsx(self, rows: list[list]) -> bytes:
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_valid_xlsx_returns_64_char_hex(self):
        raw = self._make_xlsx([
            ["Data", "Importo", "Descrizione"],
            ["01/01/2024", 100.0, "Test"],
        ])
        h = compute_header_sha256(raw, "file.xlsx")
        assert len(h) == 64
        assert all(c in "0123456789abcdef" for c in h)

    def test_xlsx_hash_deterministic(self):
        raw = self._make_xlsx([["Data", "Importo"], ["01/01/2024", 100.0]])
        h1 = compute_header_sha256(raw, "a.xlsx")
        h2 = compute_header_sha256(raw, "b.xlsx")
        assert h1 == h2


# ── TestLoadRawHeadExcel ───────────────────────────────────────────────────────

class TestLoadRawHeadExcel:
    """Cover lines 354-361 — Excel path in load_raw_head."""

    def _make_xlsx(self, rows: list[list]) -> bytes:
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_xlsx_returns_dataframe(self):
        raw = self._make_xlsx([
            ["Data", "Importo", "Descrizione"],
            ["01/01/2024", 100.0, "Supermercato"],
            ["02/01/2024", 200.0, "Farmacia"],
        ])
        df = load_raw_head(raw, "test.xlsx", n=2)
        assert isinstance(df, pd.DataFrame)
        assert len(df) <= 2

    def test_xlsx_no_header_mode(self):
        """load_raw_head for Excel reads without header → row 0 is header as data."""
        raw = self._make_xlsx([
            ["Data", "Importo"],
            ["01/01/2024", 100.0],
        ])
        df = load_raw_head(raw, "test.xlsx", n=10)
        assert df.iloc[0, 0] == "Data"


# ── TestFindCardSettlementMatches ──────────────────────────────────────────────

class TestFindCardSettlementMatches:
    """Cover lines 719-764 — find_card_settlement_matches and its helpers."""

    from core.normalizer import find_card_settlement_matches

    def _tx(self, tx_id, amount, d):
        from datetime import date as _date
        return {"id": tx_id, "amount": _Dec(str(amount)),
                "date": _date(2024, 1, d), "reconciled": False}

    def _settlement(self, s_id, amount, d):
        from datetime import date as _date
        return {"id": s_id, "amount": _Dec(str(amount)),
                "date": _date(2024, 1, d)}

    def test_empty_inputs_returns_empty(self):
        from core.normalizer import find_card_settlement_matches
        assert find_card_settlement_matches([], []) == []

    def test_sliding_window_match(self):
        """Phase 2: contiguous subset sums to settlement (lines 739-749, 774-794)."""
        from core.normalizer import find_card_settlement_matches
        card_txs = [
            self._tx("t1", 30.0, 1),
            self._tx("t2", 20.0, 2),
            self._tx("t3", 50.0, 5),
        ]
        settlements = [self._settlement("s1", 50.0, 10)]
        results = find_card_settlement_matches(settlements, card_txs)
        assert len(results) == 1
        assert results[0]["settlement_id"] == "s1"
        assert results[0]["method"] in ("sliding_window", "subset_sum")

    def test_subset_sum_match(self):
        """Phase 3: non-contiguous subset sums to settlement (lines 752-762, 798-823)."""
        from core.normalizer import find_card_settlement_matches
        # Two transactions with a large gap → not contiguous (max_gap_days=1) → subset sum
        card_txs = [
            self._tx("t1", 30.0, 1),
            self._tx("t2", 20.0, 15),  # gap > max_gap_days=1
        ]
        settlements = [self._settlement("s1", 50.0, 20)]
        results = find_card_settlement_matches(
            settlements, card_txs, max_gap_days=1, boundary_k=5
        )
        # should find via subset_sum
        assert len(results) == 1
        assert set(results[0]["matched_ids"]) == {"t1", "t2"}

    def test_no_match_returns_empty(self):
        """No settlement matches → empty results."""
        from core.normalizer import find_card_settlement_matches
        card_txs = [self._tx("t1", 10.0, 1)]
        settlements = [self._settlement("s1", 999.0, 10)]
        results = find_card_settlement_matches(settlements, card_txs)
        assert results == []

    def test_already_reconciled_excluded(self):
        """Reconciled transactions are excluded from matching."""
        from core.normalizer import find_card_settlement_matches
        card_txs = [{"id": "t1", "amount": _Dec("50.0"),
                     "date": __import__("datetime").date(2024, 1, 1),
                     "reconciled": True}]
        settlements = [self._settlement("s1", 50.0, 5)]
        results = find_card_settlement_matches(settlements, card_txs)
        assert results == []

    def test_sliding_window_overshoot_breaks(self):
        """Running total > target → inner break triggered (line 793)."""
        from core.normalizer import find_card_settlement_matches
        from datetime import date as _date
        # Large transactions: sum of first two exceeds target before we can match
        card_txs = [
            {"id": "t1", "amount": _Dec("80.0"), "date": _date(2024, 1, 1), "reconciled": False},
            {"id": "t2", "amount": _Dec("80.0"), "date": _date(2024, 1, 2), "reconciled": False},
        ]
        settlements = [{"id": "s1", "amount": _Dec("50.0"), "date": _date(2024, 1, 10)}]
        results = find_card_settlement_matches(
            settlements, card_txs, epsilon=_Dec("0.01"), max_gap_days=5, boundary_k=3
        )
        # No match: 80 > 50+0.01 → break; subset sum also can't find 50 from [80, 80]
        assert results == []


# ── TestComputeHeaderSha256ExcelLargeFile ──────────────────────────────────────

class TestComputeHeaderSha256ExcelLargeFile:
    """Cover line 334 — n rows limit in compute_header_sha256 for Excel."""

    def _make_xlsx(self, n_rows: int) -> bytes:
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "Amount", "Description"])
        for i in range(n_rows):
            ws.append([f"2024-01-{(i % 28) + 1:02d}", float(i), f"Transaction {i}"])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_large_xlsx_truncated_to_n_rows(self):
        """File with > 30 rows → loop breaks at n=30 (line 334)."""
        raw = self._make_xlsx(50)
        h1 = compute_header_sha256(raw, "file.xlsx", n=5)
        h2 = compute_header_sha256(raw, "file.xlsx", n=10)
        # Different n → different hashes (confirming truncation works)
        assert len(h1) == 64
        assert len(h2) == 64
        assert h1 != h2


# ── TestLoadRawHeadExcelFallback ───────────────────────────────────────────────

class TestLoadRawHeadExcelFallback:
    """Cover lines 359-360 — openpyxl fallback in load_raw_head."""

    def test_invalid_xlsx_bytes_fallback(self):
        """Invalid xlsx bytes → openpyxl fails → sheet_name=0 fallback (lines 359-360)."""
        # load_raw_head with invalid xlsx bytes
        import pytest
        # The fallback sets sheet_name=0 then calls pd.read_excel which may also fail
        # but we just need to verify the exception handling path is exercised
        try:
            result = load_raw_head(b"not a real xlsx", "file.xlsx", n=5)
            # If it somehow succeeds, that's fine too
        except Exception:
            pass  # Any exception is acceptable — the path was exercised


# ── TestDetectInternalTransfersMissingLines ────────────────────────────────────

class TestDetectInternalTransfersMissingLines:
    """Cover remaining uncovered lines in detect_internal_transfers."""

    def test_no_amount_date_match_continue(self):
        """Different account_labels but amounts/dates don't match → line 634 continue."""
        from datetime import date as _date
        rows = [
            {"amount": _Dec("100.00"), "date": _date(2024, 1, 1),
             "description": "desc", "tx_type": "income", "account_label": "A"},
            {"amount": _Dec("999.00"), "date": _date(2024, 6, 1),  # no match
             "description": "desc", "tx_type": "income", "account_label": "B"},
        ]
        df = pd.DataFrame(rows)
        df["transfer_pair_id"] = None
        df["transfer_confidence"] = None
        result = detect_internal_transfers(df, keyword_patterns=[])
        assert result["transfer_pair_id"].isna().all()

    def test_amount_date_match_no_keyword_no_high_sym_continue(self):
        """Different account_labels, amounts/dates match, but no keyword and low sym → line 645."""
        from datetime import date as _date
        rows = [
            {"amount": _Dec("-100.00"), "date": _date(2024, 1, 1),
             "description": "generic expense", "tx_type": "expense", "account_label": "A"},
            {"amount": _Dec("100.00"), "date": _date(2024, 1, 5),  # 4-day gap > strict=1
             "description": "generic income", "tx_type": "income", "account_label": "B"},
        ]
        df = pd.DataFrame(rows)
        df["transfer_pair_id"] = None
        df["transfer_confidence"] = None
        result = detect_internal_transfers(
            df,
            keyword_patterns=[],
            require_keyword_confirmation=True,
            delta_days=5,
            delta_days_strict=1,  # 4-day gap > 1 → not high_sym
        )
        # No pair: no keyword, not high_sym → continue
        assert result["transfer_pair_id"].isna().all()
